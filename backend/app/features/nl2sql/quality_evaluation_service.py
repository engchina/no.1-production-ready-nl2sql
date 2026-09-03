"""NL2SQL SQL生成評価の Excel 検証、非同期実行、集計、出力。"""

from __future__ import annotations

import base64
import io
import json
import logging
import socket
import threading
import time
import uuid
from collections import Counter, defaultdict
from collections.abc import Callable
from datetime import UTC, datetime, timedelta
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook  # type: ignore[import-untyped]
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE  # type: ignore[import-untyped]
from openpyxl.styles import Alignment, Font, PatternFill  # type: ignore[import-untyped]
from openpyxl.utils import get_column_letter  # type: ignore[import-untyped]

from app.settings import get_settings

from .models import AllowedObjects, Nl2SqlEngine
from .quality_evaluation_models import (
    QualityEvaluationCapabilities,
    QualityEvaluationCase,
    QualityEvaluationDeterministicAnalysis,
    QualityEvaluationEngineCapability,
    QualityEvaluationEngineSummary,
    QualityEvaluationJobPage,
    QualityEvaluationJobRecord,
    QualityEvaluationJobSummary,
    QualityEvaluationJudge,
    QualityEvaluationJudgeCapability,
    QualityEvaluationLimits,
    QualityEvaluationResult,
    QualityEvaluationResultPage,
    QualityEvaluationStatus,
    QualityEvaluationVerdict,
    job_summary,
)
from .quality_evaluation_store import (
    MemoryQualityEvaluationRepository,
    OracleQualityEvaluationRepository,
    QualityEvaluationRepository,
)
from .service import GeneratedSql, Nl2SqlService, is_select_only, nl2sql_service

logger = logging.getLogger(__name__)

_ENGINE_LABELS = {
    Nl2SqlEngine.SELECT_AI: "Select AI",
    Nl2SqlEngine.SELECT_AI_AGENT: "Select AI Agent",
    Nl2SqlEngine.ENTERPRISE_AI_DIRECT: "Enterprise AI Direct",
}
_ALLOWED_ENGINES = frozenset(_ENGINE_LABELS)
_HEADER_ALIASES = {
    "case_id": {"ケースID", "CASE_ID", "CASEID"},
    "question": {"質問", "QUESTION"},
    "expected_sql": {"期待SQL", "EXPECTED_SQL", "EXPECTEDSQL"},
}
_TERMINAL_STATUSES = {
    QualityEvaluationStatus.COMPLETED,
    QualityEvaluationStatus.COMPLETED_WITH_ERRORS,
    QualityEvaluationStatus.FAILED,
    QualityEvaluationStatus.CANCELLED,
}
_EMPTY_ROW_STOP_THRESHOLD = 100
_SQL_NO_SPACE_AROUND = frozenset("(),=<>+-*/")


class QualityEvaluationCursorError(ValueError):
    """SQL生成評価 API のページング cursor が不正な場合のエラー。"""


def _utc_now() -> str:
    return datetime.now(UTC).isoformat()


def _encode_offset(offset: int) -> str:
    return base64.urlsafe_b64encode(str(offset).encode("ascii")).decode("ascii").rstrip("=")


def _decode_offset(cursor: str | None) -> int:
    if not cursor:
        return 0
    try:
        padded = cursor + "=" * (-len(cursor) % 4)
        return max(0, int(base64.urlsafe_b64decode(padded).decode("ascii")))
    except (ValueError, UnicodeDecodeError) as exc:
        raise QualityEvaluationCursorError("カーソルが不正です。") from exc


def _safe_excel_value(value: Any) -> Any:
    if not isinstance(value, str):
        return value
    return ILLEGAL_CHARACTERS_RE.sub("", value)


def _append_safe_excel_row(sheet: Any, values: list[Any]) -> None:
    sanitized = [_safe_excel_value(value) for value in values]
    sheet.append(sanitized)
    row_number = sheet.max_row
    for column, value in enumerate(sanitized, start=1):
        if isinstance(value, str) and value.startswith("="):
            sheet.cell(row_number, column).data_type = "s"


def _quality_evaluation_limits() -> QualityEvaluationLimits:
    settings = get_settings()
    return QualityEvaluationLimits(
        max_file_bytes=settings.nl2sql_quality_evaluation_max_file_bytes,
        max_cases=settings.nl2sql_quality_evaluation_max_cases,
        max_attempts=settings.nl2sql_quality_evaluation_max_attempts,
        attempt_timeout_seconds=max(
            1.0, settings.nl2sql_quality_evaluation_attempt_timeout_seconds
        ),
    )


def _normalized_sql(sql: str) -> str:
    text = str(sql or "")
    out: list[str] = []
    index = 0
    in_single_quote = False
    in_double_quote = False
    previous_space = False
    while index < len(text):
        char = text[index]
        next_char = text[index + 1] if index + 1 < len(text) else ""
        if in_single_quote:
            out.append(char)
            if char == "'" and next_char == "'":
                out.append(next_char)
                index += 2
                continue
            if char == "'":
                in_single_quote = False
            previous_space = False
            index += 1
            continue
        if in_double_quote:
            out.append(char)
            if char == '"' and next_char == '"':
                out.append(next_char)
                index += 2
                continue
            if char == '"':
                in_double_quote = False
            previous_space = False
            index += 1
            continue
        if char == "-" and next_char == "-":
            newline = text.find("\n", index + 2)
            if newline < 0:
                break
            if out and not previous_space and out[-1] not in _SQL_NO_SPACE_AROUND:
                out.append(" ")
                previous_space = True
            index = newline + 1
            continue
        if char == "/" and next_char == "*":
            end = text.find("*/", index + 2)
            if end < 0:
                break
            if out and not previous_space and out[-1] not in _SQL_NO_SPACE_AROUND:
                out.append(" ")
                previous_space = True
            index = end + 2
            continue
        if char == "'":
            out.append(char)
            in_single_quote = True
            previous_space = False
            index += 1
            continue
        if char == '"':
            out.append(char)
            in_double_quote = True
            previous_space = False
            index += 1
            continue
        if char.isspace():
            if out and not previous_space and out[-1] not in _SQL_NO_SPACE_AROUND:
                out.append(" ")
                previous_space = True
            index += 1
            continue
        if char in _SQL_NO_SPACE_AROUND:
            while out and out[-1] == " ":
                out.pop()
            out.append(char.upper())
            previous_space = False
            index += 1
            continue
        out.append(char.upper())
        previous_space = False
        index += 1
    normalized = "".join(out).strip()
    while normalized.endswith(";"):
        normalized = normalized[:-1].rstrip()
    return normalized


class QualityEvaluationValidationError(ValueError):
    def __init__(self, errors: list[str]) -> None:
        self.errors = errors
        super().__init__("\n".join(errors))


class QualityEvaluationJobNotFoundError(ValueError):
    """SQL生成評価 job が見つからない場合のエラー。"""


class QualityEvaluationJobStateError(ValueError):
    """SQL生成評価 job の状態が要求された操作に合わない場合のエラー。"""


class _QualityEvaluationWorkerFenceLost(RuntimeError):
    """別 worker への lease 移譲や cancel により、この worker の保存権限が失われた。"""


EngineRunner = Callable[[str, Nl2SqlEngine, str], GeneratedSql | str]
JudgeRunner = Callable[
    [str, str, str, str, QualityEvaluationDeterministicAnalysis], QualityEvaluationJudge
]


class QualityEvaluationService:
    def __init__(
        self,
        nl2sql: Nl2SqlService,
        *,
        repository: QualityEvaluationRepository | None = None,
        engine_runner: EngineRunner | None = None,
        judge_runner: JudgeRunner | None = None,
        readiness_provider: (
            Callable[[str | None], dict[Nl2SqlEngine, tuple[bool, str]]] | None
        ) = None,
    ) -> None:
        self._nl2sql = nl2sql
        self._repository = repository or self._build_repository()
        self._engine_runner = engine_runner
        self._judge_runner = judge_runner
        self._readiness_provider = readiness_provider
        self._dispatch_lock = threading.Lock()
        self._active_threads: dict[str, threading.Thread] = {}

    def _build_repository(self) -> QualityEvaluationRepository:
        settings = get_settings()
        if settings.nl2sql_persistence_mode.strip().lower() == "oracle":
            return OracleQualityEvaluationRepository(
                connection_factory=self._nl2sql._oracle_adapter.connection  # noqa: SLF001
            )
        return MemoryQualityEvaluationRepository()

    @property
    def repository_mode(self) -> str:
        return self._repository.mode

    def capabilities(self, profile_id: str | None = None) -> QualityEvaluationCapabilities:
        readiness = (
            self._readiness_provider(profile_id)
            if self._readiness_provider
            else self._nl2sql.quality_evaluation_engine_readiness(profile_id=profile_id)
        )
        if self._engine_runner and not self._readiness_provider:
            readiness = {engine: (True, "") for engine in _ALLOWED_ENGINES}
        judge_ready = bool(
            self._judge_runner or self._nl2sql._enterprise_ai_client.is_configured()  # noqa: SLF001
        )
        return QualityEvaluationCapabilities(
            engines=[
                QualityEvaluationEngineCapability(
                    engine=engine,
                    label=_ENGINE_LABELS[engine],
                    available=readiness.get(engine, (False, "利用できません。"))[0],
                    reason=readiness.get(engine, (False, "利用できません。"))[1],
                )
                for engine in _ENGINE_LABELS
            ],
            judge=QualityEvaluationJudgeCapability(
                available=judge_ready,
                reason="" if judge_ready else "OCI Enterprise AI Judge が構成されていません。",
            ),
            limits=_quality_evaluation_limits(),
        )

    def template_workbook(self) -> bytes:
        workbook = Workbook()
        cases = workbook.active
        cases.title = "cases"
        cases.append(["ケースID", "質問", "期待SQL"])
        cases.append(
            [
                "CASE-001",
                "部門ごとの売上合計を取得してください",
                "SELECT department_id, SUM(amount) FROM sales GROUP BY department_id",
            ]
        )
        readme = workbook.create_sheet("記入方法", 0)
        readme.append(["NL2SQL SQL生成評価テンプレート"])
        readme.append(["必須列", "質問、期待SQL"])
        readme.append(["任意列", "ケースID（空欄時は自動付与）"])
        readme.append(["注意", "期待SQLは SELECT/WITH のみ。数式セルは使用できません。"])
        for sheet in (cases, readme):
            sheet.freeze_panes = "A2"
            for cell in sheet[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="1F4E78")
        cases.auto_filter.ref = "A1:C2"
        for column, width in {"A": 18, "B": 48, "C": 72}.items():
            cases.column_dimensions[column].width = width
        readme.column_dimensions["A"].width = 24
        readme.column_dimensions["B"].width = 76
        buffer = io.BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()

    def parse_cases(self, content: bytes, filename: str) -> list[QualityEvaluationCase]:
        limits = _quality_evaluation_limits()
        errors: list[str] = []
        if Path(filename).suffix.lower() != ".xlsx":
            errors.append(".xlsx ファイルのみアップロードできます。")
        if len(content) > limits.max_file_bytes:
            errors.append(
                "ファイルサイズが上限 "
                f"{limits.max_file_bytes // (1024 * 1024)} MiB を超えています。"
            )
        if errors:
            raise QualityEvaluationValidationError(errors)
        try:
            workbook = load_workbook(io.BytesIO(content), data_only=False, read_only=True)
        except Exception as exc:
            raise QualityEvaluationValidationError(
                ["ファイルを Excel ブックとして読み込めません。"]
            ) from exc
        sheet = next(
            (item for item in workbook.worksheets if item.title.strip().casefold() == "cases"),
            workbook.active,
        )
        header_map: dict[str, int] = {}
        header_row = next(sheet.iter_rows(min_row=1, max_row=1), ())
        for index, cell in enumerate(header_row, start=1):
            if cell.data_type == "f":
                errors.append("行 1: ヘッダーに数式は使用できません。")
                continue
            value = str(cell.value or "").strip().upper().replace(" ", "")
            for field, aliases in _HEADER_ALIASES.items():
                if value in aliases:
                    header_map[field] = index
        required_headers = (
            ("question", "質問 / QUESTION"),
            ("expected_sql", "期待SQL / EXPECTED_SQL"),
        )
        for field, label in required_headers:
            if field not in header_map:
                errors.append(f"行 1: 必須ヘッダー「{label}」がありません。")
        if errors:
            raise QualityEvaluationValidationError(errors)

        cases: list[QualityEvaluationCase] = []
        seen_case_ids: set[str] = set()
        reserved_case_ids: set[str] = set()
        case_rows: list[tuple[int, str, str, str, list[str]]] = []
        tracked_columns = [header_map["question"], header_map["expected_sql"]]
        if "case_id" in header_map:
            tracked_columns.append(header_map["case_id"])
        max_tracked_column = max(tracked_columns)
        empty_row_streak = 0
        for row_number, row in enumerate(
            sheet.iter_rows(min_row=2, max_col=max_tracked_column),
            start=2,
        ):
            cells = [row[column - 1] for column in tracked_columns if column <= len(row)]
            if all(cell.value is None or str(cell.value).strip() == "" for cell in cells):
                empty_row_streak += 1
                if empty_row_streak >= _EMPTY_ROW_STOP_THRESHOLD:
                    break
                continue
            empty_row_streak = 0
            formula_fields = [
                name
                for name, column in header_map.items()
                if column <= len(row) and row[column - 1].data_type == "f"
            ]
            question = str(row[header_map["question"] - 1].value or "").strip()
            expected_sql = str(row[header_map["expected_sql"] - 1].value or "").strip()
            case_id_value = (
                str(row[header_map["case_id"] - 1].value or "").strip()
                if "case_id" in header_map
                else ""
            )
            if case_id_value:
                reserved_case_ids.add(case_id_value)
            case_rows.append((row_number, case_id_value, question, expected_sql, formula_fields))

        auto_case_no = 1
        for row_number, case_id_value, question, expected_sql, formula_fields in case_rows:
            if formula_fields:
                errors.append(
                    f"行 {row_number}: 数式セルは使用できません（{', '.join(formula_fields)}）。"
                )
                continue
            if case_id_value:
                case_id = case_id_value
            else:
                while True:
                    candidate = f"CASE-{auto_case_no:04d}"
                    auto_case_no += 1
                    if candidate not in reserved_case_ids and candidate not in seen_case_ids:
                        case_id = candidate
                        break
            if not question:
                errors.append(f"行 {row_number}: 質問は必須です。")
            if not expected_sql:
                errors.append(f"行 {row_number}: 期待SQLは必須です。")
            elif not is_select_only(expected_sql):
                errors.append(f"行 {row_number}: 期待SQLは SELECT/WITH のみ指定できます。")
            duplicate_case_id = case_id in seen_case_ids
            if duplicate_case_id:
                errors.append(f"行 {row_number}: ケースID「{case_id}」が重複しています。")
            else:
                # 不正な行も含め、ファイル全体で Case ID の一意性を検証する。
                seen_case_ids.add(case_id)
            if question and expected_sql and is_select_only(expected_sql) and not duplicate_case_id:
                cases.append(
                    QualityEvaluationCase(
                        case_no=len(cases) + 1,
                        case_id=case_id,
                        excel_row=row_number,
                        question=question,
                        expected_sql=expected_sql,
                    )
                )
        if not cases and not errors:
            errors.append("評価ケースがありません。")
        if len(cases) > limits.max_cases:
            errors.append(f"ケース数が上限 {limits.max_cases} 件を超えています。")
        if errors:
            raise QualityEvaluationValidationError(errors)
        return cases

    def submit(
        self,
        *,
        profile_id: str,
        engines: list[Nl2SqlEngine],
        repeat_count: int,
        content: bytes,
        filename: str,
        actor_user_uuid: str = "",
    ) -> QualityEvaluationJobSummary:
        capabilities = self.capabilities(profile_id=profile_id)
        if not capabilities.judge.available:
            raise QualityEvaluationValidationError([capabilities.judge.reason])
        if not 1 <= repeat_count <= 10:
            raise QualityEvaluationValidationError(["繰り返し回数は 1〜10 で指定してください。"])
        deduplicated_engines = list(dict.fromkeys(engines))
        if not deduplicated_engines:
            raise QualityEvaluationValidationError(["実行エンジンを1つ以上選択してください。"])
        if any(engine not in _ALLOWED_ENGINES for engine in deduplicated_engines):
            raise QualityEvaluationValidationError(["未対応の実行エンジンが指定されています。"])
        readiness = {item.engine: item for item in capabilities.engines}
        unavailable = [
            readiness[engine].reason
            for engine in deduplicated_engines
            if not readiness[engine].available
        ]
        if unavailable:
            raise QualityEvaluationValidationError(unavailable)
        try:
            profile = self._nl2sql.get_profile(profile_id)
        except ValueError as exc:
            raise QualityEvaluationValidationError([str(exc)]) from exc
        cases = self.parse_cases(content, filename)
        total_attempts = len(cases) * len(deduplicated_engines) * repeat_count
        if total_attempts > capabilities.limits.max_attempts:
            raise QualityEvaluationValidationError(
                [
                    f"総試行回数 {total_attempts} が上限 "
                    f"{capabilities.limits.max_attempts} 回を超えています。"
                ]
            )
        now = _utc_now()
        job = QualityEvaluationJobRecord(
            job_id=str(uuid.uuid4()),
            profile_id=profile.id,
            profile_name=profile.name,
            profile_category=profile.category,
            engines=deduplicated_engines,
            repeat_count=repeat_count,
            cases=cases,
            total_attempts=total_attempts,
            actor_user_uuid=actor_user_uuid,
            input_filename=Path(filename).name[:255],
            created_at=now,
            attempt_timeout_seconds=capabilities.limits.attempt_timeout_seconds,
            updated_at=now,
        )
        self._repository.save_job(job)
        logger.info(
            "quality_evaluation_submitted",
            extra={
                "job_id": job.job_id,
                "profile_id": job.profile_id,
                "engine_count": len(job.engines),
                "case_count": len(job.cases),
                "repeat_count": job.repeat_count,
            },
        )
        if get_settings().nl2sql_quality_evaluation_worker_mode.strip().lower() == "inprocess":
            self._dispatch(job.job_id)
        return job_summary(job)

    def _dispatch(self, job_id: str) -> None:
        with self._dispatch_lock:
            current = self._active_threads.get(job_id)
            if current and current.is_alive():
                return
            worker = threading.Thread(
                target=self.run_job,
                kwargs={"job_id": job_id},
                name=f"nl2sql-quality-evaluation-{job_id[:8]}",
                daemon=True,
            )
            self._active_threads[job_id] = worker
            worker.start()

    def get_job(self, job_id: str) -> QualityEvaluationJobSummary:
        job = self._repository.get_job(job_id)
        if job is None:
            raise QualityEvaluationJobNotFoundError("指定されたSQL生成評価 job が見つかりません。")
        self._wake_quality_evaluation_job_if_needed(job)
        return job_summary(job)

    def peek_job(self, job_id: str) -> QualityEvaluationJobSummary:
        return job_summary(self.peek_job_record(job_id))

    def peek_job_record(self, job_id: str) -> QualityEvaluationJobRecord:
        job = self._repository.get_job(job_id)
        if job is None:
            raise QualityEvaluationJobNotFoundError("指定されたSQL生成評価 job が見つかりません。")
        return job

    def list_jobs(
        self,
        *,
        cursor: str | None,
        limit: int,
        allowed_profile_ids: set[str] | None = None,
    ) -> QualityEvaluationJobPage:
        offset = _decode_offset(cursor)
        page_size = min(max(limit, 1), 100)
        if allowed_profile_ids is not None:
            selected: list[QualityEvaluationJobRecord] = []
            matching_total = 0
            raw_offset = 0
            while True:
                jobs, raw_total = self._repository.list_jobs(offset=raw_offset, limit=500)
                for job in jobs:
                    if (job.profile_id or "default") not in allowed_profile_ids:
                        continue
                    if matching_total >= offset and len(selected) < page_size:
                        selected.append(job)
                    matching_total += 1
                raw_offset += len(jobs)
                if raw_offset >= raw_total or not jobs:
                    break
            next_offset = offset + len(selected)
            for job in selected:
                self._wake_quality_evaluation_job_if_needed(job)
            return QualityEvaluationJobPage(
                items=[job_summary(item) for item in selected],
                next_cursor=_encode_offset(next_offset) if next_offset < matching_total else None,
                total=matching_total,
            )
        jobs, total = self._repository.list_jobs(offset=offset, limit=page_size)
        next_offset = offset + len(jobs)
        for job in jobs:
            self._wake_quality_evaluation_job_if_needed(job)
        return QualityEvaluationJobPage(
            items=[job_summary(item) for item in jobs],
            next_cursor=_encode_offset(next_offset) if next_offset < total else None,
            total=total,
        )

    def list_results(
        self, *, job_id: str, cursor: str | None, limit: int
    ) -> QualityEvaluationResultPage:
        if self._repository.get_job(job_id) is None:
            raise QualityEvaluationJobNotFoundError("指定されたSQL生成評価 job が見つかりません。")
        offset = _decode_offset(cursor)
        page_size = min(max(limit, 1), 100)
        results, total = self._repository.list_results(
            job_id=job_id, offset=offset, limit=page_size
        )
        next_offset = offset + len(results)
        return QualityEvaluationResultPage(
            items=results,
            next_cursor=_encode_offset(next_offset) if next_offset < total else None,
            total=total,
        )

    def delete_job(self, job_id: str) -> QualityEvaluationJobSummary:
        job = self._repository.get_job(job_id)
        if job is None:
            raise QualityEvaluationJobNotFoundError("指定されたSQL生成評価 job が見つかりません。")
        if job.status not in _TERMINAL_STATUSES:
            raise QualityEvaluationJobStateError(
                "実行中または待機中の SQL生成評価 job は削除できません。完了後に削除してください。"
            )
        deleted = self._repository.delete_job(job_id)
        if deleted is None:
            raise QualityEvaluationJobNotFoundError("指定されたSQL生成評価 job が見つかりません。")
        return job_summary(deleted)

    def cancel_job(self, job_id: str) -> QualityEvaluationJobSummary:
        job = self._repository.get_job(job_id)
        if job is None:
            raise QualityEvaluationJobNotFoundError("指定されたSQL生成評価 job が見つかりません。")
        if job.status == QualityEvaluationStatus.CANCELLED:
            return job_summary(job)
        if job.status in _TERMINAL_STATUSES:
            raise QualityEvaluationJobStateError(
                "このSQL生成評価 job は既に完了しているため中止できません。"
            )
        now = _utc_now()
        cancelled = job.model_copy(
            update={
                "status": QualityEvaluationStatus.CANCELLED,
                "error_message": "利用者の操作で SQL生成評価 job を中止しました。",
                "current_case_id": "",
                "current_engine": None,
                "current_repetition": 0,
                "current_attempt_started_at": None,
                "heartbeat_at": now,
                "lease_expires_at": None,
                "finished_at": now,
                "updated_at": now,
            },
            deep=True,
        )
        saved = self._repository.save_job_if_active(cancelled)
        if saved is None:
            latest = self._repository.get_job(job_id)
            if latest is None:
                raise QualityEvaluationJobNotFoundError(
                    "指定されたSQL生成評価 job が見つかりません。"
                )
            if latest.status == QualityEvaluationStatus.CANCELLED:
                return job_summary(latest)
            if latest.status in _TERMINAL_STATUSES:
                raise QualityEvaluationJobStateError(
                    "このSQL生成評価 job は既に完了しているため中止できません。"
                )
            raise QualityEvaluationJobStateError(
                "SQL生成評価 job の状態が更新されたため中止できませんでした。"
            )
        logger.info(
            "quality_evaluation_cancelled",
            extra={"job_id": saved.job_id, "profile_id": saved.profile_id},
        )
        return job_summary(saved)

    def run_next_job(self, *, worker_id: str | None = None) -> bool:
        claimed = self._repository.claim_job(
            worker_id=worker_id or self._worker_id(),
            lease_seconds=self._attempt_lease_seconds(),
        )
        if claimed is None:
            return False
        self._process_claimed_job(self._record_expired_current_attempt_if_needed(claimed))
        return True

    def run_job(self, *, job_id: str, worker_id: str | None = None) -> None:
        claimed = self._repository.claim_job(
            worker_id=worker_id or self._worker_id(),
            lease_seconds=self._attempt_lease_seconds(),
            job_id=job_id,
        )
        if claimed is not None:
            self._process_claimed_job(self._record_expired_current_attempt_if_needed(claimed))

    def _worker_id(self) -> str:
        return f"{socket.gethostname()}:{threading.get_native_id()}:{uuid.uuid4().hex[:8]}"

    def _wake_quality_evaluation_job_if_needed(
        self,
        job: QualityEvaluationJobRecord,
        *,
        settings: Any | None = None,
    ) -> bool:
        settings = settings or get_settings()
        worker_mode = settings.nl2sql_quality_evaluation_worker_mode.strip().lower()
        if worker_mode == "external":
            return False
        if job.status == QualityEvaluationStatus.PENDING:
            self._dispatch(job.job_id)
            return True
        if (
            job.status == QualityEvaluationStatus.RUNNING
            and self._quality_evaluation_lease_expired(job)
        ):
            self._dispatch(job.job_id)
            return True
        return False

    @staticmethod
    def _quality_evaluation_lease_expired(job: QualityEvaluationJobRecord) -> bool:
        if not job.lease_expires_at:
            return True
        try:
            expires_at = datetime.fromisoformat(job.lease_expires_at.replace("Z", "+00:00"))
        except ValueError:
            return True
        if expires_at.tzinfo is None:
            expires_at = expires_at.replace(tzinfo=UTC)
        return expires_at <= datetime.now(UTC)

    def _attempt_timeout_seconds(self, job: QualityEvaluationJobRecord | None = None) -> float:
        if job is not None:
            return max(1.0, float(job.attempt_timeout_seconds))
        return max(1.0, float(get_settings().nl2sql_quality_evaluation_attempt_timeout_seconds))

    def _attempt_lease_seconds(self, job: QualityEvaluationJobRecord | None = None) -> float:
        settings = get_settings()
        attempt_timeout = self._attempt_timeout_seconds(job)
        configured = max(30.0, float(settings.nl2sql_quality_evaluation_lease_seconds))
        retry_attempts = max(1, max(int(settings.oci_enterprise_ai_max_retries), 0) + 1)
        retry_floor = 2.0 * attempt_timeout * retry_attempts
        return max(configured, retry_floor)

    @staticmethod
    def _worker_fence_matches(
        job: QualityEvaluationJobRecord | None, *, worker_id: str, attempt_no: int
    ) -> bool:
        return bool(
            job
            and job.status == QualityEvaluationStatus.RUNNING
            and job.worker_id == worker_id
            and job.attempt_no == attempt_no
        )

    def _log_worker_fence_lost(
        self,
        *,
        job_id: str,
        profile_id: str,
        worker_id: str,
        attempt_no: int,
        operation: str,
    ) -> None:
        logger.info(
            "quality_evaluation_worker_fence_lost",
            extra={
                "job_id": job_id,
                "profile_id": profile_id,
                "worker_id": worker_id,
                "attempt_no": attempt_no,
                "operation": operation,
            },
        )

    def _heartbeat_current_attempt(self, job: QualityEvaluationJobRecord) -> bool:
        worker_id = job.worker_id
        attempt_no = job.attempt_no
        latest = self._repository.get_job(job.job_id)
        if not self._worker_fence_matches(latest, worker_id=worker_id, attempt_no=attempt_no):
            self._log_worker_fence_lost(
                job_id=job.job_id,
                profile_id=job.profile_id,
                worker_id=worker_id,
                attempt_no=attempt_no,
                operation="heartbeat",
            )
            return False
        if latest is None:
            return False
        now = datetime.now(UTC)
        refreshed = latest.model_copy(
            update={
                "heartbeat_at": now.isoformat(),
                "lease_expires_at": (
                    now + timedelta(seconds=self._attempt_lease_seconds(latest))
                ).isoformat(),
                "updated_at": now.isoformat(),
            },
            deep=True,
        )
        saved = self._repository.save_job_if_worker_current(
            refreshed,
            worker_id=worker_id,
            attempt_no=attempt_no,
        )
        if saved is None:
            self._log_worker_fence_lost(
                job_id=job.job_id,
                profile_id=job.profile_id,
                worker_id=worker_id,
                attempt_no=attempt_no,
                operation="heartbeat_save",
            )
            return False
        return True

    @staticmethod
    def _parse_timestamp(value: str | None) -> datetime | None:
        if not value:
            return None
        try:
            parsed = datetime.fromisoformat(value.replace("Z", "+00:00"))
        except ValueError:
            return None
        if parsed.tzinfo is None:
            parsed = parsed.replace(tzinfo=UTC)
        return parsed

    @staticmethod
    def _is_timeout_exception(exc: Exception) -> bool:
        if isinstance(exc, TimeoutError):
            return True
        message = str(exc).casefold()
        return (
            "timeout" in message
            or "timed out" in message
            or "call timeout" in message
            or "dpi-1067" in message
            or "ora-01013" in message
            or "タイムアウト" in message
        )

    def _attempt_error_message(
        self,
        exc: Exception,
        *,
        engine: Nl2SqlEngine,
        stage_label: str,
        timeout_seconds: float,
    ) -> str:
        if self._is_timeout_exception(exc):
            return (
                f"{_ENGINE_LABELS.get(engine, engine.value)} の{stage_label}が "
                f"{timeout_seconds:.0f} 秒以内に完了しなかったためタイムアウトしました。"
            )
        return str(exc)[:1000]

    def _timeout_result(
        self,
        *,
        job: QualityEvaluationJobRecord,
        case: QualityEvaluationCase,
        engine: Nl2SqlEngine,
        repetition: int,
        elapsed_ms: int,
        message: str,
    ) -> QualityEvaluationResult:
        return QualityEvaluationResult(
            result_id=str(uuid.uuid4()),
            job_id=job.job_id,
            case_no=case.case_no,
            case_id=case.case_id,
            excel_row=case.excel_row,
            question=case.question,
            expected_sql=case.expected_sql,
            engine=engine,
            repetition_no=repetition,
            generated_sql="",
            normalized_sql="",
            deterministic_analysis=QualityEvaluationDeterministicAnalysis(risk_findings=[message]),
            generation_elapsed_ms=elapsed_ms,
            judge_elapsed_ms=0,
            total_elapsed_ms=elapsed_ms,
            verdict=QualityEvaluationVerdict.NOT_ANALYZED,
            judge=None,
            generation_error=message,
            judge_error="",
            created_at=_utc_now(),
        )

    def _record_expired_current_attempt_if_needed(
        self, job: QualityEvaluationJobRecord
    ) -> QualityEvaluationJobRecord:
        if (
            job.status != QualityEvaluationStatus.RUNNING
            or not job.current_case_id
            or job.current_engine is None
            or not job.current_repetition
        ):
            return job
        started_at = self._parse_timestamp(job.current_attempt_started_at)
        if started_at is None:
            return job
        timeout_seconds = self._attempt_timeout_seconds(job)
        now = datetime.now(UTC)
        if started_at + timedelta(seconds=timeout_seconds) > now:
            return job
        case = next((item for item in job.cases if item.case_id == job.current_case_id), None)
        if case is None:
            return job
        if not self._repository.has_result(
            job_id=job.job_id,
            case_no=case.case_no,
            engine=job.current_engine.value,
            repetition_no=job.current_repetition,
        ):
            elapsed_ms = max(0, round((now - started_at).total_seconds() * 1000))
            message = (
                f"{_ENGINE_LABELS.get(job.current_engine, job.current_engine.value)} の SQL 生成が "
                f"{timeout_seconds:.0f} 秒以内に完了しなかったため、"
                "worker lease の再取得時にタイムアウト結果として記録しました。"
            )
            self._repository.save_result_if_worker_current(
                self._timeout_result(
                    job=job,
                    case=case,
                    engine=job.current_engine,
                    repetition=job.current_repetition,
                    elapsed_ms=elapsed_ms,
                    message=message,
                ),
                worker_id=job.worker_id,
                attempt_no=job.attempt_no,
            )
            logger.warning(
                "quality_evaluation_attempt_timeout_recorded",
                extra={
                    "job_id": job.job_id,
                    "profile_id": job.profile_id,
                    "engine": job.current_engine.value,
                    "case_no": case.case_no,
                    "repetition_no": job.current_repetition,
                    "timeout_seconds": timeout_seconds,
                },
            )
        return self._refresh_progress(job, worker_id=job.worker_id, attempt_no=job.attempt_no)

    def _process_claimed_job(self, job: QualityEvaluationJobRecord) -> None:
        worker_id = job.worker_id
        attempt_no = job.attempt_no
        try:
            for case in job.cases:
                for engine in job.engines:
                    for repetition in range(1, job.repeat_count + 1):
                        if self._repository.has_result(
                            job_id=job.job_id,
                            case_no=case.case_no,
                            engine=engine.value,
                            repetition_no=repetition,
                        ):
                            continue
                        latest = self._repository.get_job(job.job_id)
                        if latest is None or latest.status in _TERMINAL_STATUSES:
                            return
                        if not self._worker_fence_matches(
                            latest, worker_id=worker_id, attempt_no=attempt_no
                        ):
                            return
                        now = datetime.now(UTC)
                        timeout_seconds = self._attempt_timeout_seconds(latest)
                        job = latest.model_copy(
                            update={
                                "status": QualityEvaluationStatus.RUNNING,
                                "current_case_id": case.case_id,
                                "current_engine": engine,
                                "current_repetition": repetition,
                                "current_attempt_started_at": now.isoformat(),
                                "heartbeat_at": now.isoformat(),
                                "lease_expires_at": (
                                    now + timedelta(seconds=self._attempt_lease_seconds(latest))
                                ).isoformat(),
                                "attempt_timeout_seconds": timeout_seconds,
                                "updated_at": now.isoformat(),
                            },
                            deep=True,
                        )
                        saved_job = self._repository.save_job_if_worker_current(
                            job,
                            worker_id=worker_id,
                            attempt_no=attempt_no,
                        )
                        if saved_job is None:
                            self._log_worker_fence_lost(
                                job_id=job.job_id,
                                profile_id=job.profile_id,
                                worker_id=worker_id,
                                attempt_no=attempt_no,
                                operation="attempt_start",
                            )
                            return
                        job = saved_job
                        result = self._evaluate_attempt(job, case, engine, repetition)
                        latest = self._repository.get_job(job.job_id)
                        if latest is None or latest.status in _TERMINAL_STATUSES:
                            return
                        if not self._worker_fence_matches(
                            latest, worker_id=worker_id, attempt_no=attempt_no
                        ):
                            return
                        result_saved = self._repository.save_result_if_worker_current(
                            result,
                            worker_id=worker_id,
                            attempt_no=attempt_no,
                        )
                        latest = self._repository.get_job(job.job_id)
                        if latest is None or latest.status in _TERMINAL_STATUSES:
                            return
                        if not self._worker_fence_matches(
                            latest, worker_id=worker_id, attempt_no=attempt_no
                        ):
                            return
                        if not result_saved:
                            self._log_worker_fence_lost(
                                job_id=job.job_id,
                                profile_id=job.profile_id,
                                worker_id=worker_id,
                                attempt_no=attempt_no,
                                operation="result_save",
                            )
                            return
                        if saved_job := self._refresh_progress_after_saved_result(
                            latest,
                            result,
                            worker_id=worker_id,
                            attempt_no=attempt_no,
                        ):
                            job = saved_job
            latest = self._repository.get_job(job.job_id)
            if latest is None or latest.status in _TERMINAL_STATUSES:
                return
            if not self._worker_fence_matches(latest, worker_id=worker_id, attempt_no=attempt_no):
                return
            results = self._repository.all_results(job.job_id)
            errors = sum(1 for item in results if item.generation_error or item.judge_error)
            finished = _utc_now()
            job = latest.model_copy(
                update={
                    "status": (
                        QualityEvaluationStatus.COMPLETED_WITH_ERRORS
                        if errors
                        else QualityEvaluationStatus.COMPLETED
                    ),
                    "completed_attempts": len(results),
                    "success_count": sum(item.generation_succeeded for item in results),
                    "error_count": errors,
                    "engine_summaries": self._summaries(job, results),
                    "current_case_id": "",
                    "current_engine": None,
                    "current_repetition": 0,
                    "current_attempt_started_at": None,
                    "heartbeat_at": finished,
                    "lease_expires_at": None,
                    "finished_at": finished,
                    "updated_at": finished,
                },
                deep=True,
            )
            saved_job = self._repository.save_job_if_worker_current(
                job,
                worker_id=worker_id,
                attempt_no=attempt_no,
            )
            if saved_job is None:
                self._log_worker_fence_lost(
                    job_id=job.job_id,
                    profile_id=job.profile_id,
                    worker_id=worker_id,
                    attempt_no=attempt_no,
                    operation="completion",
                )
                return
            job = saved_job
            logger.info(
                "quality_evaluation_completed",
                extra={
                    "job_id": job.job_id,
                    "profile_id": job.profile_id,
                    "completed_attempts": job.completed_attempts,
                    "error_count": job.error_count,
                    "status": job.status.value,
                },
            )
        except _QualityEvaluationWorkerFenceLost:
            return
        except Exception as exc:
            logger.exception(
                "quality_evaluation_failed",
                extra={"job_id": job.job_id, "profile_id": job.profile_id},
            )
            latest = self._repository.get_job(job.job_id)
            if latest is None or latest.status in _TERMINAL_STATUSES:
                return
            if not self._worker_fence_matches(latest, worker_id=worker_id, attempt_no=attempt_no):
                return
            # 同メソッド上部の now(datetime)と束縛を分ける(ISO 文字列)。
            now_iso = _utc_now()
            failed = latest.model_copy(
                update={
                    "status": QualityEvaluationStatus.FAILED,
                    "error_message": str(exc)[:1000],
                    "current_attempt_started_at": None,
                    "lease_expires_at": None,
                    "finished_at": now_iso,
                    "updated_at": now_iso,
                },
                deep=True,
            )
            self._repository.save_job_if_worker_current(
                failed,
                worker_id=worker_id,
                attempt_no=attempt_no,
            )

    def _refresh_progress(
        self,
        job: QualityEvaluationJobRecord,
        *,
        worker_id: str | None = None,
        attempt_no: int | None = None,
    ) -> QualityEvaluationJobRecord:
        latest = self._repository.get_job(job.job_id)
        if latest is None or latest.status in _TERMINAL_STATUSES:
            return latest or job
        results = self._repository.all_results(job.job_id)
        now = datetime.now(UTC)
        refreshed = latest.model_copy(
            update={
                "completed_attempts": len(results),
                "success_count": sum(item.generation_succeeded for item in results),
                "error_count": sum(
                    bool(item.generation_error or item.judge_error) for item in results
                ),
                "engine_summaries": self._summaries(latest, results),
                "heartbeat_at": now.isoformat(),
                "lease_expires_at": (
                    now + timedelta(seconds=self._attempt_lease_seconds(latest))
                ).isoformat(),
                "updated_at": now.isoformat(),
            },
            deep=True,
        )
        if worker_id is not None and attempt_no is not None:
            saved = self._repository.save_job_if_worker_current(
                refreshed,
                worker_id=worker_id,
                attempt_no=attempt_no,
            )
            if saved is None:
                self._log_worker_fence_lost(
                    job_id=latest.job_id,
                    profile_id=latest.profile_id,
                    worker_id=worker_id,
                    attempt_no=attempt_no,
                    operation="progress",
                )
                return latest
            return saved
        return self._repository.save_job(refreshed)

    def _refresh_progress_after_saved_result(
        self,
        job: QualityEvaluationJobRecord,
        result: QualityEvaluationResult,
        *,
        worker_id: str,
        attempt_no: int,
    ) -> QualityEvaluationJobRecord | None:
        latest = self._repository.get_job(job.job_id)
        if latest is None or latest.status in _TERMINAL_STATUSES:
            return latest
        if not self._worker_fence_matches(latest, worker_id=worker_id, attempt_no=attempt_no):
            self._log_worker_fence_lost(
                job_id=latest.job_id,
                profile_id=latest.profile_id,
                worker_id=worker_id,
                attempt_no=attempt_no,
                operation="progress",
            )
            return latest
        now = datetime.now(UTC)
        refreshed = latest.model_copy(
            update={
                "completed_attempts": min(latest.total_attempts, latest.completed_attempts + 1),
                "success_count": latest.success_count + int(result.generation_succeeded),
                "error_count": latest.error_count
                + int(bool(result.generation_error or result.judge_error)),
                "heartbeat_at": now.isoformat(),
                "lease_expires_at": (
                    now + timedelta(seconds=self._attempt_lease_seconds(latest))
                ).isoformat(),
                "updated_at": now.isoformat(),
            },
            deep=True,
        )
        saved = self._repository.save_job_if_worker_current(
            refreshed,
            worker_id=worker_id,
            attempt_no=attempt_no,
        )
        if saved is None:
            self._log_worker_fence_lost(
                job_id=latest.job_id,
                profile_id=latest.profile_id,
                worker_id=worker_id,
                attempt_no=attempt_no,
                operation="progress",
            )
        return saved

    def _evaluate_attempt(
        self,
        job: QualityEvaluationJobRecord,
        case: QualityEvaluationCase,
        engine: Nl2SqlEngine,
        repetition: int,
    ) -> QualityEvaluationResult:
        total_started = time.perf_counter()
        generated_sql = ""
        generation_error = ""
        judge_error = ""
        judge: QualityEvaluationJudge | None = None
        analysis = QualityEvaluationDeterministicAnalysis()
        timeout_seconds = self._attempt_timeout_seconds(job)
        generation_started = time.perf_counter()
        try:
            generated = (
                self._engine_runner(case.question, engine, job.profile_id)
                if self._engine_runner
                else self._nl2sql.generate_sql_strict_for_quality_evaluation(
                    question=case.question,
                    engine=engine,
                    profile_id=job.profile_id,
                    timeout_seconds=timeout_seconds,
                    max_retries=0,
                )
            )
            generated_sql = (
                generated.generated_sql if isinstance(generated, GeneratedSql) else generated
            ).strip()
            if not generated_sql:
                raise RuntimeError("選択された engine が SQL を返しませんでした。")
        except Exception as exc:
            generation_error = self._attempt_error_message(
                exc,
                engine=engine,
                stage_label="SQL 生成",
                timeout_seconds=timeout_seconds,
            )
        generation_elapsed_ms = round((time.perf_counter() - generation_started) * 1000)
        if not self._heartbeat_current_attempt(job):
            raise _QualityEvaluationWorkerFenceLost
        judge_elapsed_ms = 0
        if generated_sql:
            try:
                allowed = self._nl2sql.resolve_allowed_objects(job.profile_id, AllowedObjects())
                local = self._nl2sql.analyze_sql(
                    generated_sql,
                    allowed,
                    self._nl2sql.get_profile(job.profile_id).default_row_limit,
                    use_llm=False,
                )
                analysis = QualityEvaluationDeterministicAnalysis(
                    is_safe=local.safety.is_safe,
                    is_select_only=local.safety.is_select_only,
                    referenced_objects=local.object_names,
                    structure_summary=local.structure_summary,
                    risk_findings=local.risk_findings,
                )
            except Exception as exc:
                analysis = QualityEvaluationDeterministicAnalysis(
                    risk_findings=[f"決定論的 SQL 解析に失敗しました: {str(exc)[:500]}"]
                )
            if not self._heartbeat_current_attempt(job):
                raise _QualityEvaluationWorkerFenceLost
            judge_started = time.perf_counter()
            try:
                judge = (
                    self._judge_runner(
                        case.question,
                        case.expected_sql,
                        generated_sql,
                        job.profile_id,
                        analysis,
                    )
                    if self._judge_runner
                    else self._judge(
                        question=case.question,
                        expected_sql=case.expected_sql,
                        generated_sql=generated_sql,
                        profile_id=job.profile_id,
                        analysis=analysis,
                        timeout_seconds=timeout_seconds,
                        max_retries=0,
                    )
                )
            except Exception as exc:
                judge_error = self._attempt_error_message(
                    exc,
                    engine=engine,
                    stage_label="LLM 判定",
                    timeout_seconds=timeout_seconds,
                )
                judge = None
            judge_elapsed_ms = round((time.perf_counter() - judge_started) * 1000)
        result = QualityEvaluationResult(
            result_id=str(uuid.uuid4()),
            job_id=job.job_id,
            case_no=case.case_no,
            case_id=case.case_id,
            excel_row=case.excel_row,
            question=case.question,
            expected_sql=case.expected_sql,
            engine=engine,
            repetition_no=repetition,
            generated_sql=generated_sql,
            normalized_sql=_normalized_sql(generated_sql) if generated_sql else "",
            deterministic_analysis=analysis,
            generation_elapsed_ms=generation_elapsed_ms,
            judge_elapsed_ms=judge_elapsed_ms,
            total_elapsed_ms=round((time.perf_counter() - total_started) * 1000),
            verdict=judge.verdict if judge else QualityEvaluationVerdict.NOT_ANALYZED,
            judge=judge,
            generation_error=generation_error,
            judge_error=judge_error,
            created_at=_utc_now(),
        )
        logger.info(
            "quality_evaluation_attempt_completed",
            extra={
                "job_id": job.job_id,
                "profile_id": job.profile_id,
                "engine": engine.value,
                "case_no": case.case_no,
                "repetition_no": repetition,
                "generation_elapsed_ms": generation_elapsed_ms,
                "judge_elapsed_ms": judge_elapsed_ms,
                "generation_succeeded": result.generation_succeeded,
                "verdict": result.verdict.value,
            },
        )
        return result

    def _judge(
        self,
        *,
        question: str,
        expected_sql: str,
        generated_sql: str,
        profile_id: str,
        analysis: QualityEvaluationDeterministicAnalysis,
        timeout_seconds: float | None = None,
        max_retries: int | None = None,
    ) -> QualityEvaluationJudge:
        profile = self._nl2sql.get_profile(profile_id)
        allowed = self._nl2sql.resolve_allowed_objects(profile_id, AllowedObjects())
        catalog = self._nl2sql._generation_schema_catalog(profile, allowed)  # noqa: SLF001
        schema_context = self._nl2sql._enterprise_ai_schema_context(  # noqa: SLF001
            profile=profile,
            allowed=allowed,
            catalog=catalog,
        )
        system_prompt = (
            "あなたは Oracle SQL の生成結果評価者です。SQL を実行せず、質問に対する期待 SQL "
            "と生成 SQL の意味が等価かを判定してください。表現の違いではなく、結合、条件、集計、"
            "NULL、重複、順序、行数制限の意味を比較します。必ず JSON object だけを返し、"
            "verdict は correct / incorrect / uncertain のいずれか、confidence は 0〜1、"
            "summary は日本語、differences と risks は日本語文字列配列、"
            "correction_suggestion は日本語文字列とします。"
        )
        prompt = json.dumps(
            {
                "question": question,
                "expected_sql": expected_sql,
                "generated_sql": generated_sql,
                "deterministic_analysis": analysis.model_dump(mode="json"),
            },
            ensure_ascii=False,
        )
        raw = self._nl2sql._enterprise_ai_client.generate(  # noqa: SLF001
            prompt=prompt,
            context=schema_context,
            system_prompt=system_prompt,
            timeout_seconds=timeout_seconds,
            max_retries=max_retries,
        )
        payload = self._nl2sql._json_object_from_text(raw)  # noqa: SLF001
        return QualityEvaluationJudge.model_validate(payload)

    def _summaries(
        self, job: QualityEvaluationJobRecord, results: list[QualityEvaluationResult]
    ) -> list[QualityEvaluationEngineSummary]:
        by_engine: dict[Nl2SqlEngine, list[QualityEvaluationResult]] = defaultdict(list)
        for result in results:
            by_engine[result.engine].append(result)
        summaries: list[QualityEvaluationEngineSummary] = []
        for engine in job.engines:
            items = by_engine[engine]
            successes = [item for item in items if item.generation_succeeded]
            verdicts = Counter(item.verdict for item in items)
            by_case: dict[int, list[str]] = defaultdict(list)
            for item in successes:
                by_case[item.case_no].append(item.normalized_sql)
            consistency_values = []
            for values in by_case.values():
                counts = Counter(values)
                consistency_values.append(max(counts.values()) / len(values))
            summaries.append(
                QualityEvaluationEngineSummary(
                    engine=engine,
                    total_attempts=len(items),
                    generation_successes=len(successes),
                    generation_success_rate=(len(successes) / len(items) if items else 0.0),
                    correct=verdicts[QualityEvaluationVerdict.CORRECT],
                    incorrect=verdicts[QualityEvaluationVerdict.INCORRECT],
                    uncertain=verdicts[QualityEvaluationVerdict.UNCERTAIN],
                    not_analyzed=verdicts[QualityEvaluationVerdict.NOT_ANALYZED],
                    normalized_sql_consistency=(
                        sum(consistency_values) / len(consistency_values)
                        if consistency_values
                        else 0.0
                    ),
                    error_count=sum(
                        bool(item.generation_error or item.judge_error) for item in items
                    ),
                )
            )
        return summaries

    def results_workbook(self, job_id: str) -> tuple[str, bytes]:
        job = self._repository.get_job(job_id)
        if job is None:
            raise QualityEvaluationJobNotFoundError("指定されたSQL生成評価 job が見つかりません。")
        if job.status not in _TERMINAL_STATUSES:
            raise QualityEvaluationJobStateError(
                "評価が完了していないため Excel をダウンロードできません。"
            )
        results = self._repository.all_results(job_id)
        workbook = Workbook()
        summary = workbook.active
        summary.title = "概要"
        summary.append(["項目", "値"])
        summary_rows: list[list[Any]] = [
            ["Job ID", job.job_id],
            ["Profile", job.profile_name],
            ["状態", job.status.value],
            ["ケース数", len(job.cases)],
            ["繰り返し回数", job.repeat_count],
            ["総試行回数", job.total_attempts],
            ["注記", "LLM 判定は補助意見であり、SQL のデータベース実行結果ではありません。"],
        ]
        for row in summary_rows:
            _append_safe_excel_row(summary, row)
        summary.append([])
        summary.append(
            [
                "エンジン",
                "生成成功率",
                "correct",
                "incorrect",
                "uncertain",
                "not_analyzed",
                "SQL一致性",
                "エラー数",
            ]
        )
        for engine_summary in self._summaries(job, results):
            summary.append(
                [
                    _ENGINE_LABELS[engine_summary.engine],
                    engine_summary.generation_success_rate,
                    engine_summary.correct,
                    engine_summary.incorrect,
                    engine_summary.uncertain,
                    engine_summary.not_analyzed,
                    engine_summary.normalized_sql_consistency,
                    engine_summary.error_count,
                ]
            )
        details = workbook.create_sheet("評価結果")
        headers = [
            "ケース番号",
            "ケースID",
            "Excel行",
            "質問",
            "期待SQL",
            "エンジン",
            "繰り返し番号",
            "生成SQL",
            "正規化SQL",
            "安全",
            "SELECTのみ",
            "参照オブジェクト",
            "構造要約",
            "リスク",
            "生成時間(ms)",
            "LLM分析時間(ms)",
            "総時間(ms)",
            "LLM判定",
            "確信度",
            "LLM分析概要",
            "差分",
            "LLMリスク",
            "修正提案",
            "生成エラー",
            "LLM分析エラー",
            "人手判定",
            "人手コメント",
        ]
        details.append(headers)
        for result in results:
            judge = result.judge
            row = [
                result.case_no,
                result.case_id,
                result.excel_row,
                result.question,
                result.expected_sql,
                _ENGINE_LABELS[result.engine],
                result.repetition_no,
                result.generated_sql,
                result.normalized_sql,
                "OK" if result.deterministic_analysis.is_safe else "NG",
                "OK" if result.deterministic_analysis.is_select_only else "NG",
                "\n".join(result.deterministic_analysis.referenced_objects),
                result.deterministic_analysis.structure_summary,
                "\n".join(result.deterministic_analysis.risk_findings),
                result.generation_elapsed_ms,
                result.judge_elapsed_ms,
                result.total_elapsed_ms,
                result.verdict.value,
                judge.confidence if judge else None,
                judge.summary if judge else "",
                "\n".join(judge.differences) if judge else "",
                "\n".join(judge.risks) if judge else "",
                judge.correction_suggestion if judge else "",
                result.generation_error,
                result.judge_error,
                "",
                "",
            ]
            _append_safe_excel_row(details, row)
        self._format_workbook(summary, details, len(headers))
        buffer = io.BytesIO()
        workbook.save(buffer)
        timestamp = datetime.now(UTC).strftime("%Y%m%d_%H%M%S")
        filename = f"nl2sql_quality_evaluation_{timestamp}_{job.job_id}.xlsx"
        return filename, buffer.getvalue()

    def _format_workbook(self, summary: Any, details: Any, detail_columns: int) -> None:
        header_fill = PatternFill("solid", fgColor="1F4E78")
        for cell in summary[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
        for cell in summary[10]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
        summary.freeze_panes = "A2"
        summary.column_dimensions["A"].width = 28
        summary.column_dimensions["B"].width = 88
        for row in summary.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for cell in details[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        details.freeze_panes = "A2"
        details.auto_filter.ref = f"A1:{get_column_letter(detail_columns)}{max(details.max_row, 1)}"
        widths = [12, 18, 10, 42, 64, 24, 12, 64, 64, 10, 12, 32, 38, 38]
        widths += [14, 16, 14, 16, 12, 40, 40, 40, 48, 38, 38, 16, 40]
        for index, width in enumerate(widths, start=1):
            details.column_dimensions[get_column_letter(index)].width = width
        for row in details.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)


quality_evaluation_service = QualityEvaluationService(nl2sql_service)
