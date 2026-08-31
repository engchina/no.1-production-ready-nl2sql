from __future__ import annotations

import io
from datetime import UTC, datetime, timedelta
from typing import Any

import pytest
from openpyxl import Workbook, load_workbook  # type: ignore[import-untyped]

from app.features.nl2sql.models import (
    AssetRefreshData,
    Nl2SqlEngine,
    Nl2SqlProfile,
    SampleDataMutationRequest,
    SampleDataStep,
)
from app.features.nl2sql.quality_evaluation_models import (
    QualityEvaluationCase,
    QualityEvaluationJobRecord,
    QualityEvaluationJudge,
    QualityEvaluationResult,
    QualityEvaluationStatus,
    QualityEvaluationVerdict,
)
from app.features.nl2sql.quality_evaluation_service import (
    QualityEvaluationJobStateError,
    QualityEvaluationService,
    QualityEvaluationValidationError,
)
from app.features.nl2sql.quality_evaluation_store import (
    MemoryQualityEvaluationRepository,
    OracleQualityEvaluationRepository,
)
from app.features.nl2sql.service import Nl2SqlService
from app.features.nl2sql.store import MemoryNl2SqlStore
from app.main import app
from app.settings import get_settings


def _xlsx(
    rows: list[list[object]],
    *,
    headers: list[str] | None = None,
    sheet_name: str = "cases",
) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    sheet.append(headers or ["ケースID", "質問", "期待SQL"])
    for row in rows:
        sheet.append(row)
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _service(**kwargs: object) -> QualityEvaluationService:
    return QualityEvaluationService(
        Nl2SqlService(store=MemoryNl2SqlStore()),
        repository=MemoryQualityEvaluationRepository(),
        **kwargs,  # type: ignore[arg-type]
    )


def _create_sample_profile(service: Nl2SqlService) -> Nl2SqlProfile:
    return service.create_profile(
        Nl2SqlProfile(
            id="sql_assist_sample",
            name="SQL_ASSIST_SAMPLE",
            category="SAMPLE",
            description="SQL Assist sample objects test profile.",
            allowed_tables=["DEPARTMENT", "EMPLOYEE", "PROJECT"],
            allowed_views=["V_EMP_DEPT", "V_DEPT_PROJECT"],
        )
    )


def _judge(*_args: object) -> QualityEvaluationJudge:
    return QualityEvaluationJudge(
        verdict=QualityEvaluationVerdict.CORRECT,
        confidence=0.92,
        summary="質問の意味と一致します。",
    )


def _evaluation_job(
    job_id: str,
    *,
    status: QualityEvaluationStatus = QualityEvaluationStatus.PENDING,
    created_at: datetime | None = None,
    lease_expires_at: str | None = None,
) -> QualityEvaluationJobRecord:
    now = created_at or datetime.now(UTC)
    finished_at = (
        now.isoformat()
        if status
        in {
            QualityEvaluationStatus.COMPLETED,
            QualityEvaluationStatus.COMPLETED_WITH_ERRORS,
            QualityEvaluationStatus.FAILED,
            QualityEvaluationStatus.CANCELLED,
        }
        else None
    )
    return QualityEvaluationJobRecord(
        job_id=job_id,
        profile_id="default",
        profile_name="default",
        engines=[Nl2SqlEngine.SELECT_AI],
        repeat_count=1,
        cases=[
            QualityEvaluationCase(
                case_no=1,
                case_id="A",
                excel_row=2,
                question="q",
                expected_sql="SELECT 1 FROM dual",
            )
        ],
        status=status,
        total_attempts=1,
        lease_expires_at=lease_expires_at,
        created_at=now.isoformat(),
        finished_at=finished_at,
        updated_at=now.isoformat(),
    )


def _evaluation_result(job_id: str) -> QualityEvaluationResult:
    now = datetime.now(UTC).isoformat()
    return QualityEvaluationResult(
        result_id=f"{job_id}-result-1",
        job_id=job_id,
        case_no=1,
        case_id="A",
        excel_row=2,
        question="q",
        expected_sql="SELECT 1 FROM dual",
        engine=Nl2SqlEngine.SELECT_AI,
        repetition_no=1,
        generated_sql="SELECT 1 FROM dual",
        normalized_sql="SELECT 1 FROM DUAL",
        verdict=QualityEvaluationVerdict.CORRECT,
        created_at=now,
    )


class _FakeEnterpriseAiClient:
    def __init__(self, text: str) -> None:
        self.text = text
        self.calls: list[dict[str, Any]] = []

    def is_configured(self) -> bool:
        return True

    def model_id(self) -> str:
        return "enterprise-nl2sql-model"

    def generate(
        self,
        *,
        prompt: str,
        context: str,
        system_prompt: str,
        timeout_seconds: float | None = None,
    ) -> str:
        self.calls.append(
            {
                "prompt": prompt,
                "context": context,
                "system_prompt": system_prompt,
                "timeout_seconds": timeout_seconds,
            }
        )
        return self.text


def _configure_strict_engine_runtime(
    service: Nl2SqlService, monkeypatch: pytest.MonkeyPatch
) -> None:
    settings = get_settings()
    monkeypatch.setattr(settings, "nl2sql_select_ai_enabled", True)
    monkeypatch.setattr(settings, "nl2sql_select_ai_agent_enabled", True)
    monkeypatch.setattr(settings, "nl2sql_enterprise_ai_direct_enabled", True)
    monkeypatch.setattr(settings, "nl2sql_select_ai_provider", "oci")
    monkeypatch.setattr(settings, "nl2sql_select_ai_credential_name", "NL2SQL_CRED")
    monkeypatch.setattr(service, "_use_oracle_runtime", lambda: True)
    monkeypatch.setattr(service._oracle_adapter, "is_configured", lambda: True)
    monkeypatch.setattr(service._enterprise_ai_client, "is_configured", lambda: True)


class _LobPayload:
    def __init__(self, value: str) -> None:
        self.value = value


class _ConnectionBoundLob:
    def __init__(self, connection: _QualityEvaluationOracleConnection, value: str) -> None:
        self._connection = connection
        self._value = value

    def read(self) -> str:
        if self._connection.closed:
            raise RuntimeError("LOB locator is no longer valid")
        return self._value


_ResultSet = list[tuple[object, ...]]


class _QualityEvaluationOracleCursor:
    def __init__(
        self,
        connection: _QualityEvaluationOracleConnection,
        result_sets: list[_ResultSet],
    ) -> None:
        self._connection = connection
        self._result_sets = result_sets
        self._current: _ResultSet = []
        self.executed: list[tuple[str, Any]] = []

    def __enter__(self) -> _QualityEvaluationOracleCursor:
        return self

    def __exit__(self, *_args: object) -> None:
        return None

    def execute(self, sql: str, binds: Any = None) -> None:
        self.executed.append((sql, binds))
        if not sql.lstrip().upper().startswith("SELECT"):
            self._current = []
            return
        if not self._result_sets:
            raise AssertionError("scripted result set is missing")
        self._current = [self._materialize_row(row) for row in self._result_sets.pop(0)]

    def fetchone(self) -> tuple[object, ...] | None:
        return self._current[0] if self._current else None

    def fetchall(self) -> _ResultSet:
        return list(self._current)

    def _materialize_row(self, row: tuple[object, ...]) -> tuple[object, ...]:
        return tuple(
            (
                _ConnectionBoundLob(self._connection, value.value)
                if isinstance(value, _LobPayload)
                else value
            )
            for value in row
        )


class _QualityEvaluationOracleConnection:
    def __init__(self, result_sets: list[_ResultSet]) -> None:
        self._result_sets = result_sets
        self.closed = False
        self.cursors: list[_QualityEvaluationOracleCursor] = []
        self.commits = 0
        self.rollbacks = 0

    def __enter__(self) -> _QualityEvaluationOracleConnection:
        return self

    def __exit__(self, *_args: object) -> None:
        self.closed = True

    def cursor(self) -> _QualityEvaluationOracleCursor:
        cursor = _QualityEvaluationOracleCursor(self, self._result_sets)
        self.cursors.append(cursor)
        return cursor

    def commit(self) -> None:
        self.commits += 1

    def rollback(self) -> None:
        self.rollbacks += 1


class _QualityEvaluationOracleConnectionFactory:
    def __init__(self, calls: list[list[_ResultSet]]) -> None:
        self._calls = calls
        self.connections: list[_QualityEvaluationOracleConnection] = []

    def __call__(self) -> _QualityEvaluationOracleConnection:
        if not self._calls:
            raise AssertionError("scripted connection is missing")
        connection = _QualityEvaluationOracleConnection(self._calls.pop(0))
        self.connections.append(connection)
        return connection


def test_parse_cases_accepts_japanese_and_english_headers_and_active_sheet() -> None:
    service = _service()
    japanese = service.parse_cases(
        _xlsx([["JP-1", "一覧を取得", "SELECT * FROM orders"]]), "cases.xlsx"
    )
    english = service.parse_cases(
        _xlsx(
            [["EN-1", "count rows", "WITH q AS (SELECT 1 x FROM dual) SELECT * FROM q"]],
            headers=["CASE_ID", "QUESTION", "EXPECTED_SQL"],
            sheet_name="Sheet1",
        ),
        "cases.xlsx",
    )
    assert japanese[0].case_id == "JP-1"
    assert english[0].excel_row == 2


def test_parse_cases_ignores_empty_rows_and_assigns_case_id() -> None:
    cases = _service().parse_cases(
        _xlsx([[None, None, None], [None, "件数", "SELECT COUNT(*) FROM orders"]]),
        "cases.xlsx",
    )
    assert len(cases) == 1
    assert cases[0].case_id == "CASE-0001"
    assert cases[0].excel_row == 3


@pytest.mark.parametrize(
    ("rows", "expected"),
    [
        ([["A", "質問", "=A1"]], "数式セル"),
        (
            [["A", "質問1", "SELECT 1 FROM dual"], ["A", "質問2", "SELECT 2 FROM dual"]],
            "重複",
        ),
        ([["A", "質問", "DELETE FROM orders"]], "SELECT/WITH"),
        ([["A", "", "SELECT 1 FROM dual"]], "質問は必須"),
    ],
)
def test_parse_cases_rejects_entire_workbook_with_excel_row_errors(
    rows: list[list[object]], expected: str
) -> None:
    with pytest.raises(QualityEvaluationValidationError) as caught:
        _service().parse_cases(_xlsx(rows), "cases.xlsx")
    assert expected in str(caught.value)
    assert "行 " in str(caught.value)


def test_parse_cases_rejects_extension_and_file_limit(monkeypatch: pytest.MonkeyPatch) -> None:
    settings = get_settings()
    monkeypatch.setattr(settings, "nl2sql_quality_evaluation_max_file_bytes", 32)
    with pytest.raises(QualityEvaluationValidationError) as caught:
        _service().parse_cases(b"x" * 33, "cases.xls")
    assert ".xlsx" in str(caught.value)
    assert "ファイルサイズ" in str(caught.value)


def test_parse_cases_enforces_case_limit(monkeypatch: pytest.MonkeyPatch) -> None:
    settings = get_settings()
    monkeypatch.setattr(settings, "nl2sql_quality_evaluation_max_cases", 1)
    with pytest.raises(QualityEvaluationValidationError, match="ケース数が上限 1"):
        _service().parse_cases(
            _xlsx(
                [
                    ["A", "質問1", "SELECT 1 FROM dual"],
                    ["B", "質問2", "SELECT 2 FROM dual"],
                ]
            ),
            "cases.xlsx",
        )


def test_submit_validates_engines_repeat_profile_and_attempt_limit(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    service = _service(engine_runner=lambda *_args: "SELECT 1 FROM dual", judge_runner=_judge)
    workbook = _xlsx([["A", "質問", "SELECT 1 FROM dual"]])
    with pytest.raises(QualityEvaluationValidationError, match="1つ以上"):
        service.submit(
            profile_id="default",
            engines=[],
            repeat_count=1,
            content=workbook,
            filename="cases.xlsx",
        )
    with pytest.raises(QualityEvaluationValidationError, match="1〜10"):
        service.submit(
            profile_id="default",
            engines=[Nl2SqlEngine.SELECT_AI],
            repeat_count=11,
            content=workbook,
            filename="cases.xlsx",
        )
    with pytest.raises(QualityEvaluationValidationError, match="profile"):
        service.submit(
            profile_id="missing",
            engines=[Nl2SqlEngine.SELECT_AI],
            repeat_count=1,
            content=workbook,
            filename="cases.xlsx",
        )
    settings = get_settings()
    monkeypatch.setattr(settings, "nl2sql_quality_evaluation_max_attempts", 1)
    with pytest.raises(QualityEvaluationValidationError, match="総試行回数"):
        service.submit(
            profile_id="default",
            engines=[Nl2SqlEngine.SELECT_AI, Nl2SqlEngine.SELECT_AI_AGENT],
            repeat_count=1,
            content=workbook,
            filename="cases.xlsx",
        )


def test_submit_accepts_repeat_boundaries_and_rejects_unavailable_engine(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    settings = get_settings()
    monkeypatch.setattr(settings, "nl2sql_quality_evaluation_worker_mode", "external")
    workbook = _xlsx([["A", "質問", "SELECT 1 FROM dual"]])
    service = _service(engine_runner=lambda *_args: "SELECT 1 FROM dual", judge_runner=_judge)
    assert (
        service.submit(
            profile_id="default",
            engines=[Nl2SqlEngine.SELECT_AI],
            repeat_count=1,
            content=workbook,
            filename="cases.xlsx",
        ).total_attempts
        == 1
    )
    assert (
        service.submit(
            profile_id="default",
            engines=[Nl2SqlEngine.SELECT_AI],
            repeat_count=10,
            content=workbook,
            filename="cases.xlsx",
        ).total_attempts
        == 10
    )

    unavailable = _service(
        engine_runner=lambda *_args: "SELECT 1 FROM dual",
        judge_runner=_judge,
        readiness_provider=lambda _profile_id: {
            Nl2SqlEngine.SELECT_AI: (False, "Select AI profile が未構成です。")
        },
    )
    with pytest.raises(QualityEvaluationValidationError, match="未構成"):
        unavailable.submit(
            profile_id="default",
            engines=[Nl2SqlEngine.SELECT_AI],
            repeat_count=1,
            content=workbook,
            filename="cases.xlsx",
        )


def test_quality_evaluation_get_wakes_orphan_pending_job_inprocess(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    settings = get_settings()
    monkeypatch.setattr(settings, "nl2sql_quality_evaluation_worker_mode", "inprocess")
    repository = MemoryQualityEvaluationRepository()
    repository.save_job(_evaluation_job("pending-job"))
    service = QualityEvaluationService(
        Nl2SqlService(store=MemoryNl2SqlStore()), repository=repository
    )
    dispatched: list[str] = []
    monkeypatch.setattr(service, "_dispatch", lambda job_id: dispatched.append(job_id))

    fetched = service.get_job("pending-job")

    assert fetched.status == QualityEvaluationStatus.PENDING
    assert dispatched == ["pending-job"]


def test_quality_evaluation_get_wakes_expired_running_job_only_inprocess(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    settings = get_settings()
    monkeypatch.setattr(settings, "nl2sql_quality_evaluation_worker_mode", "inprocess")
    now = datetime.now(UTC)
    repository = MemoryQualityEvaluationRepository()
    repository.save_job(
        _evaluation_job(
            "expired-running",
            status=QualityEvaluationStatus.RUNNING,
            lease_expires_at=(now - timedelta(seconds=1)).isoformat(),
        )
    )
    repository.save_job(
        _evaluation_job(
            "active-running",
            status=QualityEvaluationStatus.RUNNING,
            lease_expires_at=(now + timedelta(minutes=5)).isoformat(),
        )
    )
    repository.save_job(_evaluation_job("completed", status=QualityEvaluationStatus.COMPLETED))
    service = QualityEvaluationService(
        Nl2SqlService(store=MemoryNl2SqlStore()), repository=repository
    )
    dispatched: list[str] = []
    monkeypatch.setattr(service, "_dispatch", lambda job_id: dispatched.append(job_id))

    service.get_job("expired-running")
    service.get_job("active-running")
    service.get_job("completed")

    assert dispatched == ["expired-running"]


def test_quality_evaluation_get_does_not_wake_external_worker_jobs(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    settings = get_settings()
    monkeypatch.setattr(settings, "nl2sql_quality_evaluation_worker_mode", "external")
    repository = MemoryQualityEvaluationRepository()
    repository.save_job(_evaluation_job("pending-job"))
    service = QualityEvaluationService(
        Nl2SqlService(store=MemoryNl2SqlStore()), repository=repository
    )
    dispatched: list[str] = []
    monkeypatch.setattr(service, "_dispatch", lambda job_id: dispatched.append(job_id))

    fetched = service.get_job("pending-job")

    assert fetched.status == QualityEvaluationStatus.PENDING
    assert dispatched == []


def test_quality_evaluation_list_wakes_visible_orphan_jobs_inprocess(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    settings = get_settings()
    monkeypatch.setattr(settings, "nl2sql_quality_evaluation_worker_mode", "inprocess")
    now = datetime.now(UTC)
    repository = MemoryQualityEvaluationRepository()
    repository.save_job(_evaluation_job("older-pending", created_at=now - timedelta(minutes=2)))
    repository.save_job(_evaluation_job("newer-pending", created_at=now - timedelta(minutes=1)))
    repository.save_job(_evaluation_job("completed", status=QualityEvaluationStatus.COMPLETED))
    service = QualityEvaluationService(
        Nl2SqlService(store=MemoryNl2SqlStore()), repository=repository
    )
    dispatched: list[str] = []
    monkeypatch.setattr(service, "_dispatch", lambda job_id: dispatched.append(job_id))

    page = service.list_jobs(cursor=None, limit=10)

    assert [item.job_id for item in page.items] == [
        "completed",
        "newer-pending",
        "older-pending",
    ]
    assert dispatched == ["newer-pending", "older-pending"]


def test_quality_evaluation_delete_terminal_job_removes_job_and_results() -> None:
    repository = MemoryQualityEvaluationRepository()
    repository.save_job(_evaluation_job("completed-job", status=QualityEvaluationStatus.COMPLETED))
    repository.save_result(_evaluation_result("completed-job"))
    service = QualityEvaluationService(
        Nl2SqlService(store=MemoryNl2SqlStore()), repository=repository
    )

    deleted = service.delete_job("completed-job")

    assert deleted.job_id == "completed-job"
    assert repository.get_job("completed-job") is None
    results, total = repository.list_results(job_id="completed-job", offset=0, limit=10)
    assert results == []
    assert total == 0


def test_quality_evaluation_delete_missing_job_raises_not_found() -> None:
    service = _service()

    with pytest.raises(ValueError, match="見つかりません"):
        service.delete_job("missing-job")


@pytest.mark.parametrize(
    "status",
    [QualityEvaluationStatus.PENDING, QualityEvaluationStatus.RUNNING],
)
def test_quality_evaluation_delete_rejects_active_jobs(status: QualityEvaluationStatus) -> None:
    repository = MemoryQualityEvaluationRepository()
    repository.save_job(_evaluation_job("active-job", status=status))
    repository.save_result(_evaluation_result("active-job"))
    service = QualityEvaluationService(
        Nl2SqlService(store=MemoryNl2SqlStore()), repository=repository
    )

    with pytest.raises(QualityEvaluationJobStateError, match="削除できません"):
        service.delete_job("active-job")

    assert repository.get_job("active-job") is not None
    results, total = repository.list_results(job_id="active-job", offset=0, limit=10)
    assert [item.result_id for item in results] == ["active-job-result-1"]
    assert total == 1


def test_quality_evaluation_capabilities_exposes_attempt_timeout(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    settings = get_settings()
    monkeypatch.setattr(settings, "nl2sql_quality_evaluation_attempt_timeout_seconds", 300.0)
    service = _service(engine_runner=lambda *_args: "SELECT 1 FROM dual", judge_runner=_judge)

    capabilities = service.capabilities(profile_id="default")

    assert capabilities.limits.attempt_timeout_seconds == 300.0


def test_quality_evaluation_cancel_is_idempotent_and_rejects_completed_jobs() -> None:
    repository = MemoryQualityEvaluationRepository()
    repository.save_job(_evaluation_job("pending-job"))
    repository.save_job(_evaluation_job("finished-job", status=QualityEvaluationStatus.COMPLETED))
    service = QualityEvaluationService(
        Nl2SqlService(store=MemoryNl2SqlStore()), repository=repository
    )

    cancelled = service.cancel_job("pending-job")
    cancelled_again = service.cancel_job("pending-job")

    assert cancelled.status == QualityEvaluationStatus.CANCELLED
    assert cancelled_again.status == QualityEvaluationStatus.CANCELLED
    assert cancelled_again.finished_at is not None
    with pytest.raises(QualityEvaluationJobStateError, match="中止できません"):
        service.cancel_job("finished-job")
    with pytest.raises(ValueError, match="見つかりません"):
        service.cancel_job("missing-job")


def test_worker_records_select_ai_agent_timeout_as_error_result(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    settings = get_settings()
    monkeypatch.setattr(settings, "nl2sql_quality_evaluation_worker_mode", "external")
    monkeypatch.setattr(settings, "nl2sql_quality_evaluation_attempt_timeout_seconds", 300.0)

    def engine(_question: str, selected: Nl2SqlEngine, _profile: str) -> str:
        assert selected == Nl2SqlEngine.SELECT_AI_AGENT
        raise RuntimeError("DPI-1067: call timeout of 300000 ms exceeded")

    service = _service(engine_runner=engine, judge_runner=_judge)
    submitted = service.submit(
        profile_id="default",
        engines=[Nl2SqlEngine.SELECT_AI_AGENT],
        repeat_count=1,
        content=_xlsx([["A", "質問", "SELECT 1 FROM dual"]]),
        filename="cases.xlsx",
    )

    service.run_job(job_id=submitted.job_id)

    result = service.list_results(job_id=submitted.job_id, cursor=None, limit=1).items[0]
    assert result.verdict == QualityEvaluationVerdict.NOT_ANALYZED
    assert "Select AI Agent" in result.generation_error
    assert "300 秒" in result.generation_error
    assert "タイムアウト" in result.generation_error
    assert service.get_job(submitted.job_id).status == QualityEvaluationStatus.COMPLETED_WITH_ERRORS


def test_expired_lease_reclaim_synthesizes_timeout_result(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    settings = get_settings()
    monkeypatch.setattr(settings, "nl2sql_quality_evaluation_worker_mode", "external")
    now = datetime.now(UTC)
    repository = MemoryQualityEvaluationRepository()
    repository.save_job(
        QualityEvaluationJobRecord(
            job_id="expired-attempt",
            profile_id="default",
            profile_name="default",
            engines=[Nl2SqlEngine.SELECT_AI_AGENT],
            repeat_count=1,
            cases=[
                QualityEvaluationCase(
                    case_no=1,
                    case_id="A",
                    excel_row=2,
                    question="q",
                    expected_sql="SELECT 1 FROM dual",
                )
            ],
            status=QualityEvaluationStatus.RUNNING,
            total_attempts=1,
            current_case_id="A",
            current_engine=Nl2SqlEngine.SELECT_AI_AGENT,
            current_repetition=1,
            current_attempt_started_at=(now - timedelta(seconds=360)).isoformat(),
            heartbeat_at=(now - timedelta(seconds=360)).isoformat(),
            lease_expires_at=(now - timedelta(seconds=1)).isoformat(),
            attempt_timeout_seconds=300.0,
            created_at=now.isoformat(),
            updated_at=now.isoformat(),
        )
    )

    def engine(*_args: object) -> str:
        raise AssertionError("expired attempt should not be executed again")

    service = QualityEvaluationService(
        Nl2SqlService(store=MemoryNl2SqlStore()),
        repository=repository,
        engine_runner=engine,
        judge_runner=_judge,
    )

    service.run_job(job_id="expired-attempt", worker_id="reclaimer")

    result = service.list_results(job_id="expired-attempt", cursor=None, limit=1).items[0]
    job = service.get_job("expired-attempt")
    assert "worker lease" in result.generation_error
    assert "タイムアウト" in result.generation_error
    assert result.verdict == QualityEvaluationVerdict.NOT_ANALYZED
    assert job.status == QualityEvaluationStatus.COMPLETED_WITH_ERRORS
    assert job.completed_attempts == 1


def test_cancelled_job_is_not_overwritten_by_late_worker_return(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    settings = get_settings()
    monkeypatch.setattr(settings, "nl2sql_quality_evaluation_worker_mode", "external")
    service = _service(judge_runner=_judge)
    submitted_job_id = ""

    def engine(_question: str, _selected: Nl2SqlEngine, _profile: str) -> str:
        service.cancel_job(submitted_job_id)
        return "SELECT 1 FROM dual"

    service._engine_runner = engine  # noqa: SLF001
    submitted = service.submit(
        profile_id="default",
        engines=[Nl2SqlEngine.SELECT_AI],
        repeat_count=1,
        content=_xlsx([["A", "質問", "SELECT 1 FROM dual"]]),
        filename="cases.xlsx",
    )
    submitted_job_id = submitted.job_id

    service.run_job(job_id=submitted.job_id)

    job = service.get_job(submitted.job_id)
    results = service.list_results(job_id=submitted.job_id, cursor=None, limit=10)
    assert job.status == QualityEvaluationStatus.CANCELLED
    assert job.completed_attempts == 0
    assert results.items == []


def test_quality_evaluation_readiness_allows_profile_engines_without_cached_asset_meta(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    service = Nl2SqlService(store=MemoryNl2SqlStore())
    _configure_strict_engine_runtime(service, monkeypatch)

    readiness = service.quality_evaluation_engine_readiness(profile_id="default")

    assert readiness[Nl2SqlEngine.SELECT_AI] == (True, "")
    assert readiness[Nl2SqlEngine.SELECT_AI_AGENT] == (True, "")
    assert readiness[Nl2SqlEngine.ENTERPRISE_AI_DIRECT] == (True, "")


def test_quality_evaluation_readiness_blocks_known_unsynced_current_profile(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    service = Nl2SqlService(store=MemoryNl2SqlStore())
    _configure_strict_engine_runtime(service, monkeypatch)
    profile_name = service._select_ai_profile_name(service.get_profile("default"))  # noqa: SLF001
    service._asset_meta[Nl2SqlEngine.SELECT_AI] = AssetRefreshData(  # noqa: SLF001
        engine=Nl2SqlEngine.SELECT_AI,
        refreshed=False,
        status="error",
        profile_name=profile_name,
        engine_meta={
            "profile_scope_states": {profile_name.upper(): {"refreshed": False, "status": "error"}}
        },
    )

    readiness = service.quality_evaluation_engine_readiness(profile_id="default")

    assert readiness[Nl2SqlEngine.SELECT_AI][0] is False
    assert "未同期" in readiness[Nl2SqlEngine.SELECT_AI][1]
    assert readiness[Nl2SqlEngine.SELECT_AI_AGENT][0] is False
    assert "未同期" in readiness[Nl2SqlEngine.SELECT_AI_AGENT][1]


def test_quality_evaluation_readiness_ignores_other_profile_stale_asset_meta(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    service = Nl2SqlService(store=MemoryNl2SqlStore())
    _configure_strict_engine_runtime(service, monkeypatch)
    service._asset_meta[Nl2SqlEngine.SELECT_AI] = AssetRefreshData(  # noqa: SLF001
        engine=Nl2SqlEngine.SELECT_AI,
        refreshed=False,
        status="error",
        profile_name="NL2SQL_OTHER_PROFILE",
    )
    service._asset_meta[Nl2SqlEngine.SELECT_AI_AGENT] = AssetRefreshData(  # noqa: SLF001
        engine=Nl2SqlEngine.SELECT_AI_AGENT,
        refreshed=False,
        status="error",
        profile_name="NL2SQL_OTHER_PROFILE",
    )

    readiness = service.quality_evaluation_engine_readiness(profile_id="default")

    assert readiness[Nl2SqlEngine.SELECT_AI] == (True, "")
    assert readiness[Nl2SqlEngine.SELECT_AI_AGENT] == (True, "")


def test_worker_runs_every_attempt_and_isolates_generation_and_judge_errors(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    calls: dict[tuple[str, Nl2SqlEngine], int] = {}

    def engine(question: str, selected: Nl2SqlEngine, _profile: str) -> str:
        key = (question, selected)
        calls[key] = calls.get(key, 0) + 1
        if selected == Nl2SqlEngine.SELECT_AI_AGENT and calls[key] == 1:
            raise RuntimeError("strict engine failure")
        return "SELECT 1 FROM dual"

    def judge(
        question: str,
        _expected: str,
        _generated: str,
        _profile: str,
        _analysis: object,
    ) -> QualityEvaluationJudge:
        if question == "judge failure":
            raise RuntimeError("invalid judge response")
        return _judge()

    settings = get_settings()
    monkeypatch.setattr(settings, "nl2sql_quality_evaluation_worker_mode", "external")
    service = _service(engine_runner=engine, judge_runner=judge)
    submitted = service.submit(
        profile_id="default",
        engines=[Nl2SqlEngine.SELECT_AI, Nl2SqlEngine.SELECT_AI_AGENT],
        repeat_count=2,
        content=_xlsx(
            [
                ["A", "ok", "SELECT 1 FROM dual"],
                ["B", "judge failure", "SELECT 1 FROM dual"],
            ]
        ),
        filename="cases.xlsx",
    )
    service.run_job(job_id=submitted.job_id, worker_id="test-worker")
    job = service.get_job(submitted.job_id)
    page = service.list_results(job_id=submitted.job_id, cursor=None, limit=100)
    assert job.status == QualityEvaluationStatus.COMPLETED_WITH_ERRORS
    assert job.completed_attempts == 8
    assert len(page.items) == 8
    assert sum(bool(item.generation_error) for item in page.items) == 2
    assert sum(bool(item.judge_error) for item in page.items) == 3
    assert sum(item.verdict == QualityEvaluationVerdict.CORRECT for item in page.items) == 3


def test_worker_aggregates_verdict_distribution_and_normalized_consistency(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    settings = get_settings()
    monkeypatch.setattr(settings, "nl2sql_quality_evaluation_worker_mode", "external")
    call_count = 0

    def engine(*_args: object) -> str:
        nonlocal call_count
        call_count += 1
        return "SELECT 1 FROM dual" if call_count == 1 else "SELECT 2 FROM dual"

    def judge(
        _question: str,
        _expected: str,
        generated: str,
        _profile: str,
        _analysis: object,
    ) -> QualityEvaluationJudge:
        return QualityEvaluationJudge(
            verdict=(
                QualityEvaluationVerdict.CORRECT
                if "SELECT 1" in generated
                else QualityEvaluationVerdict.INCORRECT
            ),
            confidence=0.8,
            summary="比較しました。",
        )

    service = _service(engine_runner=engine, judge_runner=judge)
    submitted = service.submit(
        profile_id="default",
        engines=[Nl2SqlEngine.SELECT_AI],
        repeat_count=2,
        content=_xlsx([["A", "質問", "SELECT 1 FROM dual"]]),
        filename="cases.xlsx",
    )
    service.run_job(job_id=submitted.job_id)
    summary = service.get_job(submitted.job_id).engine_summaries[0]
    assert summary.generation_success_rate == 1.0
    assert summary.correct == 1
    assert summary.incorrect == 1
    assert summary.normalized_sql_consistency == 0.5


def test_worker_records_empty_engine_output_as_generation_error(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    settings = get_settings()
    monkeypatch.setattr(settings, "nl2sql_quality_evaluation_worker_mode", "external")
    service = _service(engine_runner=lambda *_args: "   ", judge_runner=_judge)
    submitted = service.submit(
        profile_id="default",
        engines=[Nl2SqlEngine.SELECT_AI],
        repeat_count=1,
        content=_xlsx([["A", "質問", "SELECT 1 FROM dual"]]),
        filename="cases.xlsx",
    )
    service.run_job(job_id=submitted.job_id)
    result = service.list_results(job_id=submitted.job_id, cursor=None, limit=1).items[0]
    assert result.generated_sql == ""
    assert "SQL を返しませんでした" in result.generation_error
    assert result.verdict == QualityEvaluationVerdict.NOT_ANALYZED
    assert service.get_job(submitted.job_id).status == QualityEvaluationStatus.COMPLETED_WITH_ERRORS


def test_default_judge_uses_profile_schema_catalog_without_argument_error(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    settings = get_settings()
    monkeypatch.setattr(settings, "nl2sql_quality_evaluation_worker_mode", "external")
    nl2sql = Nl2SqlService(store=MemoryNl2SqlStore())
    nl2sql.import_sample_data(
        SampleDataMutationRequest(step=SampleDataStep.ALL, confirmation="SQL_ASSIST_SAMPLE")
    )
    _create_sample_profile(nl2sql)
    fake_client = _FakeEnterpriseAiClient(
        '{"verdict":"correct","confidence":0.91,"summary":"意味が一致します。",'
        '"differences":[],"risks":[],"correction_suggestion":""}'
    )
    nl2sql._enterprise_ai_client = fake_client  # noqa: SLF001
    service = QualityEvaluationService(
        nl2sql,
        repository=MemoryQualityEvaluationRepository(),
        engine_runner=lambda *_args: "SELECT EMPLOYEE_ID FROM EMPLOYEE",
    )

    submitted = service.submit(
        profile_id="sql_assist_sample",
        engines=[Nl2SqlEngine.ENTERPRISE_AI_DIRECT],
        repeat_count=1,
        content=_xlsx([["A", "社員IDを取得してください", "SELECT EMPLOYEE_ID FROM EMPLOYEE"]]),
        filename="cases.xlsx",
    )
    service.run_job(job_id=submitted.job_id)

    result = service.list_results(job_id=submitted.job_id, cursor=None, limit=1).items[0]
    assert result.judge_error == ""
    assert "_enterprise_ai_schema_context" not in result.judge_error
    assert result.verdict == QualityEvaluationVerdict.CORRECT
    assert fake_client.calls
    assert fake_client.calls[0]["timeout_seconds"] == 300.0
    assert "schema:" in fake_client.calls[0]["context"]
    assert "EMPLOYEE" in fake_client.calls[0]["context"]


def test_strict_enterprise_ai_direct_generation_passes_timeout_override(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    service = Nl2SqlService(store=MemoryNl2SqlStore())
    service.import_sample_data(
        SampleDataMutationRequest(step=SampleDataStep.ALL, confirmation="SQL_ASSIST_SAMPLE")
    )
    _create_sample_profile(service)
    fake_client = _FakeEnterpriseAiClient(
        '{"sql":"SELECT EMPLOYEE_ID FROM EMPLOYEE","explanation":"社員IDを取得します。"}'
    )
    service._enterprise_ai_client = fake_client  # noqa: SLF001
    monkeypatch.setattr(
        service,
        "quality_evaluation_engine_readiness",
        lambda profile_id=None: {Nl2SqlEngine.ENTERPRISE_AI_DIRECT: (True, "")},
    )
    monkeypatch.setattr(service, "_use_oracle_runtime", lambda: False)

    generated = service.generate_sql_strict_for_quality_evaluation(
        question="社員IDを取得してください",
        engine=Nl2SqlEngine.ENTERPRISE_AI_DIRECT,
        profile_id="sql_assist_sample",
        timeout_seconds=123.0,
    )

    assert generated.generated_sql == "SELECT EMPLOYEE_ID FROM EMPLOYEE"
    assert fake_client.calls[0]["timeout_seconds"] == 123.0


@pytest.mark.parametrize(
    "engine",
    [
        Nl2SqlEngine.SELECT_AI,
        Nl2SqlEngine.SELECT_AI_AGENT,
        Nl2SqlEngine.ENTERPRISE_AI_DIRECT,
    ],
)
def test_strict_generation_never_uses_deterministic_fallback(
    engine: Nl2SqlEngine, monkeypatch: pytest.MonkeyPatch
) -> None:
    service = Nl2SqlService(store=MemoryNl2SqlStore())
    service.import_sample_data(
        SampleDataMutationRequest(step=SampleDataStep.ALL, confirmation="SQL_ASSIST_SAMPLE")
    )
    monkeypatch.setattr(
        service,
        "quality_evaluation_engine_readiness",
        lambda profile_id=None: {engine: (True, "")},
    )
    monkeypatch.setattr(service, "_use_oracle_runtime", lambda: False)
    monkeypatch.setattr(service._enterprise_ai_client, "is_configured", lambda: False)
    with pytest.raises(RuntimeError):
        service.generate_sql_strict_for_quality_evaluation(
            question="社員一覧",
            engine=engine,
            profile_id="default",
        )


def test_memory_repository_reclaims_expired_lease_and_preserves_result_idempotency() -> None:
    repository = MemoryQualityEvaluationRepository()
    now = datetime.now(UTC)
    job = QualityEvaluationJobRecord(
        job_id="job-1",
        profile_id="default",
        profile_name="default",
        engines=[Nl2SqlEngine.SELECT_AI],
        repeat_count=1,
        cases=[
            QualityEvaluationCase(
                case_no=1,
                case_id="A",
                excel_row=2,
                question="q",
                expected_sql="SELECT 1 FROM dual",
            )
        ],
        status=QualityEvaluationStatus.RUNNING,
        total_attempts=1,
        lease_expires_at=(now - timedelta(seconds=1)).isoformat(),
        created_at=now.isoformat(),
        updated_at=now.isoformat(),
    )
    repository.save_job(job)
    reclaimed = repository.claim_job(worker_id="new-worker", lease_seconds=60)
    assert reclaimed is not None
    assert reclaimed.worker_id == "new-worker"
    assert reclaimed.attempt_no == 1


def test_oracle_repository_materializes_quality_evaluation_lobs_before_connection_closes(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    configured_cursors: list[_QualityEvaluationOracleCursor] = []

    def configure_clob_fetch(cursor: _QualityEvaluationOracleCursor) -> None:
        configured_cursors.append(cursor)

    monkeypatch.setattr(
        "app.features.nl2sql.quality_evaluation_store.configure_clob_fetch_as_text",
        configure_clob_fetch,
    )
    now = datetime.now(UTC).isoformat()
    job = QualityEvaluationJobRecord(
        job_id="job-1",
        profile_id="default",
        profile_name="default",
        engines=[Nl2SqlEngine.SELECT_AI],
        repeat_count=1,
        cases=[
            QualityEvaluationCase(
                case_no=1,
                case_id="A",
                excel_row=2,
                question="q",
                expected_sql="SELECT 1 FROM dual",
            )
        ],
        total_attempts=1,
        created_at=now,
        updated_at=now,
    )
    result = QualityEvaluationResult(
        result_id="result-1",
        job_id=job.job_id,
        case_no=1,
        case_id="A",
        excel_row=2,
        question="q",
        expected_sql="SELECT 1 FROM dual",
        engine=Nl2SqlEngine.SELECT_AI,
        repetition_no=1,
        generated_sql="SELECT 1 FROM dual",
        normalized_sql="SELECT 1 FROM DUAL",
        verdict=QualityEvaluationVerdict.CORRECT,
        created_at=now,
    )
    job_payload = job.model_dump_json()
    result_payload = result.model_dump_json()
    factory = _QualityEvaluationOracleConnectionFactory(
        [
            [[(_LobPayload(job_payload),)]],
            [[(1,)], [(_LobPayload(job_payload),)]],
            [[(1,)], [(_LobPayload(result_payload),)]],
            [[(_LobPayload(result_payload),)]],
            [[("job-1", _LobPayload(job_payload))]],
        ]
    )
    repository = OracleQualityEvaluationRepository(connection_factory=factory)

    fetched = repository.get_job("job-1")
    jobs, total_jobs = repository.list_jobs(offset=0, limit=10)
    results, total_results = repository.list_results(job_id="job-1", offset=0, limit=10)
    all_results = repository.all_results("job-1")
    claimed = repository.claim_job(worker_id="worker-1", lease_seconds=60, job_id="job-1")

    assert fetched is not None
    assert fetched.job_id == "job-1"
    assert total_jobs == 1
    assert [item.job_id for item in jobs] == ["job-1"]
    assert total_results == 1
    assert [item.result_id for item in results] == ["result-1"]
    assert [item.result_id for item in all_results] == ["result-1"]
    assert claimed is not None
    assert claimed.worker_id == "worker-1"
    assert claimed.status == QualityEvaluationStatus.RUNNING
    assert len(configured_cursors) == 5
    claim_cursor = factory.connections[-1].cursors[0]
    claim_sql = claim_cursor.executed[0][0].upper()
    assert "FOR UPDATE SKIP LOCKED" in claim_sql
    assert "FETCH FIRST" not in claim_sql
    assert claim_cursor.prefetchrows == 0
    assert claim_cursor.arraysize == 1
    assert all(connection.closed for connection in factory.connections)


def test_oracle_repository_delete_job_uses_job_delete_and_transaction() -> None:
    job = _evaluation_job("job-1", status=QualityEvaluationStatus.COMPLETED)
    factory = _QualityEvaluationOracleConnectionFactory(
        [
            [[(_LobPayload(job.model_dump_json()),)]],
        ]
    )
    repository = OracleQualityEvaluationRepository(connection_factory=factory)

    deleted = repository.delete_job("job-1")

    assert deleted is not None
    assert deleted.job_id == "job-1"
    connection = factory.connections[0]
    assert connection.commits == 1
    assert connection.rollbacks == 0
    executed_sql = [sql for cursor in connection.cursors for sql, _binds in cursor.executed]
    assert any(
        "DELETE FROM NL2SQL_EVALUATION_JOBS WHERE JOB_ID = :job_id" in sql for sql in executed_sql
    )


def test_oracle_repository_delete_job_rolls_back_on_delete_error() -> None:
    class _FailingDeleteCursor(_QualityEvaluationOracleCursor):
        def execute(self, sql: str, binds: Any = None) -> None:
            if sql.lstrip().upper().startswith("DELETE"):
                self.executed.append((sql, binds))
                raise RuntimeError("delete failed")
            super().execute(sql, binds)

    class _FailingDeleteConnection(_QualityEvaluationOracleConnection):
        def cursor(self) -> _QualityEvaluationOracleCursor:
            cursor = _FailingDeleteCursor(self, self._result_sets)
            self.cursors.append(cursor)
            return cursor

    class _FailingDeleteConnectionFactory(_QualityEvaluationOracleConnectionFactory):
        def __call__(self) -> _QualityEvaluationOracleConnection:
            if not self._calls:
                raise AssertionError("scripted connection is missing")
            connection = _FailingDeleteConnection(self._calls.pop(0))
            self.connections.append(connection)
            return connection

    job = _evaluation_job("job-1", status=QualityEvaluationStatus.COMPLETED)
    factory = _FailingDeleteConnectionFactory(
        [
            [[(_LobPayload(job.model_dump_json()),)]],
        ]
    )
    repository = OracleQualityEvaluationRepository(connection_factory=factory)

    with pytest.raises(RuntimeError, match="delete failed"):
        repository.delete_job("job-1")

    connection = factory.connections[0]
    assert connection.commits == 0
    assert connection.rollbacks == 1


def test_results_workbook_has_review_columns_format_and_formula_injection_guard(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    settings = get_settings()
    monkeypatch.setattr(settings, "nl2sql_quality_evaluation_worker_mode", "external")
    service = _service(
        engine_runner=lambda *_args: "SELECT 1 FROM dual",
        judge_runner=_judge,
    )
    submitted = service.submit(
        profile_id="default",
        engines=[Nl2SqlEngine.ENTERPRISE_AI_DIRECT],
        repeat_count=1,
        content=_xlsx([["+CASE", "+question", "SELECT 1 FROM dual"]]),
        filename="cases.xlsx",
    )
    service.run_job(job_id=submitted.job_id)
    filename, content = service.results_workbook(submitted.job_id)
    workbook = load_workbook(io.BytesIO(content))
    details = workbook["評価結果"]
    headers = [cell.value for cell in details[1]]
    assert filename.startswith("nl2sql_quality_evaluation_")
    assert workbook.sheetnames == ["概要", "評価結果"]
    assert headers[-2:] == ["人手判定", "人手コメント"]
    assert details.freeze_panes == "A2"
    assert details.auto_filter.ref is not None
    assert str(details.cell(2, 2).value).startswith("'")
    assert str(details.cell(2, 4).value).startswith("'")


def test_result_and_job_pagination_use_opaque_cursors(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    settings = get_settings()
    monkeypatch.setattr(settings, "nl2sql_quality_evaluation_worker_mode", "external")
    service = _service(
        engine_runner=lambda *_args: "SELECT 1 FROM dual",
        judge_runner=_judge,
    )
    first = service.submit(
        profile_id="default",
        engines=[Nl2SqlEngine.SELECT_AI],
        repeat_count=2,
        content=_xlsx([["A", "質問", "SELECT 1 FROM dual"]]),
        filename="cases.xlsx",
    )
    service.submit(
        profile_id="default",
        engines=[Nl2SqlEngine.SELECT_AI],
        repeat_count=1,
        content=_xlsx([["B", "質問", "SELECT 1 FROM dual"]]),
        filename="cases.xlsx",
    )
    service.run_job(job_id=first.job_id)

    jobs_page = service.list_jobs(cursor=None, limit=1)
    results_page = service.list_results(job_id=first.job_id, cursor=None, limit=1)
    assert len(jobs_page.items) == 1
    assert jobs_page.next_cursor
    assert len(service.list_jobs(cursor=jobs_page.next_cursor, limit=1).items) == 1
    assert len(results_page.items) == 1
    assert results_page.next_cursor
    assert (
        len(
            service.list_results(
                job_id=first.job_id,
                cursor=results_page.next_cursor,
                limit=1,
            ).items
        )
        == 1
    )


def test_openapi_exposes_new_quality_flow_and_removes_legacy_evaluation_routes() -> None:
    paths = app.openapi()["paths"]
    expected = {
        "/api/nl2sql/quality-evaluations/capabilities",
        "/api/nl2sql/quality-evaluations/template.xlsx",
        "/api/nl2sql/quality-evaluations",
        "/api/nl2sql/quality-evaluations/{job_id}",
        "/api/nl2sql/quality-evaluations/{job_id}/cancel",
        "/api/nl2sql/quality-evaluations/{job_id}/results",
        "/api/nl2sql/quality-evaluations/{job_id}/results.xlsx",
    }
    assert expected <= set(paths)
    assert "post" in paths["/api/nl2sql/quality-evaluations/{job_id}/cancel"]
    assert "delete" in paths["/api/nl2sql/quality-evaluations/{job_id}"]
    assert "/api/nl2sql/analyze" in paths
    assert "/api/nl2sql/repair" not in paths
    assert "/api/nl2sql/reverse" in paths
    assert "/api/nl2sql/synthetic-data/generate" in paths
    removed = {
        "/api/nl2sql/evaluate",
        "/api/nl2sql/evaluation-sets",
        "/api/nl2sql/evaluation-runs",
        "/api/nl2sql/compare",
        "/api/nl2sql/synthetic-cases",
        "/api/nl2sql/synthetic-data/operations/{operation_id}",
    }
    assert removed.isdisjoint(paths)
