"""改善・運用系 NL2SQL API の profile アクセス境界の回帰テスト。"""

from __future__ import annotations

import io
import json
from datetime import UTC, datetime
from types import SimpleNamespace
from typing import cast

import pytest
from fastapi import HTTPException, Request

from app.features.nl2sql import ontology_router
from app.features.nl2sql import profile_sync as profile_sync_module
from app.features.nl2sql import router as nl2sql_router
from app.features.nl2sql.models import (
    AdminFeedbackReviewRequest,
    AssetCleanupRequest,
    ClassifierPredictRequest,
    ClassifierTrainingExampleUpdateRequest,
    ClassifierTrainRequest,
    FeedbackEntriesDeleteRequest,
    FeedbackIndexRequest,
    FeedbackRating,
    HistoryItem,
    Nl2SqlEngine,
    Nl2SqlProfile,
    ProfileRecommendationRequest,
    ProfileSyncJobData,
    ProfileSyncJobStatus,
)
from app.features.nl2sql.ontology_models import (
    ProfileRecommendation,
    ProfileRecommendationCandidateV2,
    utc_now,
)
from app.features.nl2sql.quality_evaluation_models import (
    QualityEvaluationJobPage,
    QualityEvaluationJobSummary,
    QualityEvaluationResultPage,
    QualityEvaluationStatus,
)
from app.features.nl2sql.quality_evaluation_service import QualityEvaluationJobNotFoundError
from app.features.nl2sql.service import Nl2SqlService
from app.features.nl2sql.store import MemoryNl2SqlStore
from app.security.domain import Principal


class _DisabledEmbeddingClient:
    def is_configured(self) -> bool:
        return False

    def embed_texts(self, _texts: list[str]) -> list[list[float]]:
        raise AssertionError("deterministic fallback should be used")


def _principal(allowed_profile_ids: set[str]) -> Principal:
    return Principal(
        user_uuid="user-1",
        login_user_id="user1",
        display_name="利用者1",
        status="ACTIVE",
        force_password_change=False,
        role_codes=["GENERAL"],
        permissions=set(),
        data_entitlements=[],
        allowed_profile_ids=allowed_profile_ids,
        session_id="session-1",
        csrf_token_hash="hash",
        password_change_allowed=True,
    )


def _request(allowed_profile_ids: set[str]) -> Request:
    return cast(
        Request, SimpleNamespace(state=SimpleNamespace(principal=_principal(allowed_profile_ids)))
    )


def _quality_job(job_id: str, profile_id: str) -> QualityEvaluationJobSummary:
    now = datetime(2026, 9, 3, tzinfo=UTC).isoformat()
    return QualityEvaluationJobSummary(
        job_id=job_id,
        profile_id=profile_id,
        profile_name=f"{profile_id} profile",
        engines=[Nl2SqlEngine.ENTERPRISE_AI_DIRECT],
        repeat_count=1,
        case_count=1,
        total_attempts=1,
        completed_attempts=0,
        success_count=0,
        error_count=0,
        status=QualityEvaluationStatus.PENDING,
        created_at=now,
        started_at=None,
        finished_at=None,
        updated_at=now,
    )


class _QualityEvaluationService:
    def __init__(self) -> None:
        self.jobs = {
            "job-a": _quality_job("job-a", "profile-a"),
            "job-b": _quality_job("job-b", "profile-b"),
        }
        self.job_actors = {"job-a": "user-1", "job-b": "user-1"}
        self.cancelled: list[str] = []
        self.deleted: list[str] = []
        self.woken: list[str] = []

    def peek_job(self, job_id: str) -> QualityEvaluationJobSummary:
        job = self.jobs.get(job_id)
        if job is None:
            raise QualityEvaluationJobNotFoundError("指定されたSQL生成評価 job が見つかりません。")
        return job

    def peek_job_record(self, job_id: str) -> SimpleNamespace:
        job = self.peek_job(job_id)
        return SimpleNamespace(
            job_id=job.job_id,
            profile_id=job.profile_id,
            actor_user_uuid=self.job_actors.get(job_id, ""),
        )

    def get_job(self, job_id: str) -> QualityEvaluationJobSummary:
        self.woken.append(job_id)
        return self.peek_job(job_id)

    def list_jobs(
        self,
        *,
        cursor: str | None,
        limit: int,
        allowed_profile_ids: set[str] | None = None,
    ) -> QualityEvaluationJobPage:
        del cursor, limit
        items = list(self.jobs.values())
        if allowed_profile_ids is not None:
            items = [item for item in items if item.profile_id in allowed_profile_ids]
        self.woken.extend(item.job_id for item in items)
        return QualityEvaluationJobPage(items=items, total=len(items))

    def list_results(
        self, *, job_id: str, cursor: str | None, limit: int
    ) -> QualityEvaluationResultPage:
        del job_id, cursor, limit
        return QualityEvaluationResultPage(items=[], total=0)

    def results_workbook(self, job_id: str) -> tuple[str, bytes]:
        return f"{job_id}.xlsx", b"workbook"

    def cancel_job(self, job_id: str) -> QualityEvaluationJobSummary:
        self.cancelled.append(job_id)
        return self.peek_job(job_id)

    def delete_job(self, job_id: str) -> QualityEvaluationJobSummary:
        self.deleted.append(job_id)
        return self.peek_job(job_id)


class _ProfileSyncService:
    def __init__(self) -> None:
        self.retried: list[str] = []
        self.jobs = {
            "sync-b": ProfileSyncJobData(
                job_id="sync-b",
                profile_id="profile-b",
                status=ProfileSyncJobStatus.FAILED,
            )
        }

    def get(self, job_id: str) -> ProfileSyncJobData | None:
        return self.jobs.get(job_id)

    def retry(self, job_id: str) -> ProfileSyncJobData:
        self.retried.append(job_id)
        return self.jobs[job_id]


def _classifier_artifact(*classes: str) -> bytes:
    vector = [0.0] * 1536
    payload = {
        "classes": list(classes),
        "coef": [vector],
        "intercept": [5.0],
        "feature_dim": 1536,
        "embedding_model": "deterministic-hash-1536",
    }
    return json.dumps(payload, ensure_ascii=False).encode("utf-8")


def _training_xlsx(rows: list[tuple[str, str]]) -> bytes:
    import openpyxl  # type: ignore[import-untyped]

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "ClassifierTraining"
    sheet.append(["PROFILE_ID", "TEXT"])
    for profile_id, text in rows:
        sheet.append([profile_id, text])
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _history_item(history_id: str, profile_id: str) -> HistoryItem:
    return HistoryItem(
        id=history_id,
        question=f"{profile_id} の質問",
        engine=Nl2SqlEngine.SELECT_AI,
        generated_sql="SELECT 1 FROM DUAL",
        created_at=datetime(2026, 9, 3, tzinfo=UTC).isoformat(),
        feedback_rating=FeedbackRating.GOOD,
        profile_id=profile_id,
        profile_name=f"{profile_id} profile",
        admin_feedback_rating=FeedbackRating.GOOD,
        admin_feedback_content="承認済み",
    )


def test_quality_evaluation_routes_enforce_profile_access(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    service = _QualityEvaluationService()
    monkeypatch.setattr(nl2sql_router, "quality_evaluation_service", service)
    request = _request({"profile-a"})

    listed = nl2sql_router.list_quality_evaluations(request)
    assert [item.profile_id for item in listed.data.items] == ["profile-a"]  # type: ignore[union-attr]
    assert listed.data.total == 1  # type: ignore[union-attr]

    for call in (
        lambda: nl2sql_router.get_quality_evaluation("job-b", request),
        lambda: nl2sql_router.quality_evaluation_results("job-b", request),
        lambda: nl2sql_router.quality_evaluation_results_xlsx("job-b", request),
        lambda: nl2sql_router.cancel_quality_evaluation("job-b", request),
        lambda: nl2sql_router.delete_quality_evaluation("job-b", request),
    ):
        with pytest.raises(HTTPException) as exc_info:
            call()
        assert exc_info.value.status_code == 403

    assert service.cancelled == []
    assert service.deleted == []
    assert service.woken == ["job-a"]


def test_quality_evaluation_cancel_delete_enforce_actor_owner(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    service = _QualityEvaluationService()
    service.job_actors["job-a"] = "other-user"
    monkeypatch.setattr(nl2sql_router, "quality_evaluation_service", service)
    request = _request({"profile-a"})

    with pytest.raises(HTTPException) as cancel_exc:
        nl2sql_router.cancel_quality_evaluation("job-a", request)
    with pytest.raises(HTTPException) as delete_exc:
        nl2sql_router.delete_quality_evaluation("job-a", request)

    assert cancel_exc.value.status_code == 403
    assert delete_exc.value.status_code == 403
    assert service.cancelled == []
    assert service.deleted == []
    assert service.woken == []


def test_feedback_learning_routes_enforce_profile_access(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    service = Nl2SqlService(store=MemoryNl2SqlStore())
    with service._lock:  # noqa: SLF001
        service._history = [  # noqa: SLF001
            _history_item("history-a", "profile-a"),
            _history_item("history-b", "profile-b"),
        ]
        service._feedback_indexed_ids = {"history-a", "history-b"}  # noqa: SLF001
    monkeypatch.setattr(nl2sql_router, "nl2sql_service", service)
    request = _request({"profile-a"})

    feedback = nl2sql_router.list_feedback(request)
    assert [item.profile_id for item in feedback.data.items] == ["profile-a"]  # type: ignore[union-attr]

    entries = nl2sql_router.feedback_entries(request)
    assert [item.profile_id for item in entries.data.items] == ["profile-a"]  # type: ignore[union-attr]
    assert entries.data.indexed_count == 1  # type: ignore[union-attr]

    index_status = nl2sql_router.feedback_index_status(request)
    assert index_status.data.source_history_count == 1  # type: ignore[union-attr]
    assert index_status.data.indexable_count == 1  # type: ignore[union-attr]
    assert index_status.data.indexed_count == 1  # type: ignore[union-attr]

    for call in (
        lambda: nl2sql_router.admin_review_feedback(
            AdminFeedbackReviewRequest(
                history_id="history-b",
                rating=FeedbackRating.GOOD,
                feedback_content="承認",
            ),
            request,
        ),
        lambda: nl2sql_router.delete_feedback_entries(
            FeedbackEntriesDeleteRequest(history_ids=["history-b"]),
            request,
        ),
    ):
        with pytest.raises(HTTPException) as exc_info:
            call()
        assert exc_info.value.status_code == 403

    nl2sql_router.rebuild_feedback_index(FeedbackIndexRequest(), request)
    nl2sql_router.clear_feedback_index(FeedbackIndexRequest(), request)
    assert service.list_feedback_entries().total == 2


def test_select_ai_refresh_and_cleanup_require_default_profile_access(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    service = SimpleNamespace(
        refresh_select_ai_profile=lambda _profile_id=None: pytest.fail(
            "refresh should be blocked before service call"
        ),
        refresh_select_ai_agent_assets=lambda _profile_id=None: pytest.fail(
            "agent refresh should be blocked before service call"
        ),
        cleanup_select_ai_assets=lambda **_kwargs: pytest.fail(
            "cleanup should be blocked before service call"
        ),
        cleanup_select_ai_agent_assets_low_level=lambda _req: pytest.fail(
            "agent cleanup should be blocked before service call"
        ),
    )
    monkeypatch.setattr(nl2sql_router, "nl2sql_service", service)
    request = _request({"profile-a"})

    for call in (
        lambda: nl2sql_router.refresh_select_ai_profile(request),
        lambda: nl2sql_router.refresh_select_ai_agent_assets(request),
        lambda: nl2sql_router.cleanup_select_ai_assets(AssetCleanupRequest(), request),
        lambda: nl2sql_router.cleanup_select_ai_agent_assets(AssetCleanupRequest(), request),
    ):
        with pytest.raises(HTTPException) as exc_info:
            call()
        assert exc_info.value.status_code == 403


def test_classifier_training_import_delete_and_scores_enforce_profile_access(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    service = Nl2SqlService(store=MemoryNl2SqlStore())
    service._embedding_client = _DisabledEmbeddingClient()  # type: ignore[assignment]  # noqa: SLF001
    service.create_profile(
        Nl2SqlProfile(id="profile-a", name="許可プロファイル", allowed_tables=["APP.A"])
    )
    service.create_profile(
        Nl2SqlProfile(id="profile-b", name="拒否プロファイル", allowed_tables=["APP.B"])
    )
    service.create_profile(
        Nl2SqlProfile(id="profile-c", name="追加許可プロファイル", allowed_tables=["APP.C"])
    )
    imported = service.import_classifier_training_data(
        filename="training.xlsx",
        content=_training_xlsx(
            [
                ("profile-a", "許可された質問"),
                ("profile-b", "許可外の質問"),
            ]
        ),
        replace=True,
        allowed_profile_ids={"profile-a"},
    )
    assert imported.imported_count == 1
    assert imported.skipped_count == 1
    assert [item.profile_id for item in service.classifier_training_data().examples] == [
        "profile-a"
    ]
    assert any("利用権限がない Profile" in warning for warning in imported.warnings)

    unrestricted = service.import_classifier_training_data(
        filename="training.xlsx",
        content=_training_xlsx(
            [
                ("profile-b", "削除対象の質問"),
                ("profile-c", "追加許可の質問"),
            ]
        ),
    )
    denied_example_id = next(
        item.id for item in unrestricted.examples if item.profile_id == "profile-b"
    )
    monkeypatch.setattr(nl2sql_router, "nl2sql_service", service)
    listed = nl2sql_router.classifier_training_data(_request({"profile-a"}))
    assert [item.profile_id for item in listed.data.examples] == ["profile-a"]  # type: ignore[union-attr]

    response = nl2sql_router.export_classifier_training_data_xlsx(_request({"profile-a"}))
    import openpyxl

    workbook = openpyxl.load_workbook(io.BytesIO(response.body), read_only=True)
    sheet = workbook.active
    exported_profile_ids = [
        row[2]
        for row in sheet.iter_rows(min_row=2, values_only=True)
        if row and any(value is not None for value in row)
    ]
    assert exported_profile_ids == ["profile-a"]

    with pytest.raises(HTTPException) as patch_exc:
        nl2sql_router.update_classifier_training_example(
            denied_example_id,
            ClassifierTrainingExampleUpdateRequest(
                text="許可プロファイルへ移動したい質問",
                profile_id="profile-a",
            ),
            _request({"profile-a"}),
        )
    assert patch_exc.value.status_code == 403

    with pytest.raises(HTTPException) as exc_info:
        nl2sql_router.delete_classifier_training_example(
            denied_example_id,
            _request({"profile-a"}),
        )
    assert exc_info.value.status_code == 403

    replaced = service.import_classifier_training_data(
        filename="training.xlsx",
        content=_training_xlsx([("profile-a", "置換後の質問")]),
        replace=True,
        allowed_profile_ids={"profile-a"},
    )
    assert replaced.imported_count == 1
    all_examples = service.classifier_training_data().examples
    assert any(item.profile_id == "profile-b" for item in all_examples)
    assert any(item.profile_id == "profile-c" for item in all_examples)
    assert any(item.text == "置換後の質問" for item in all_examples)
    assert all(item.text != "許可された質問" for item in all_examples)

    trained_status = nl2sql_router.train_classifier(
        ClassifierTrainRequest(),
        _request({"profile-a", "profile-c"}),
    )
    assert set(trained_status.data.categories) == {"profile-a", "profile-c"}  # type: ignore[union-attr]
    with service._lock:  # noqa: SLF001
        trained_categories = set(
            (service._classifier_artifact or {}).get("categories", [])
        )  # noqa: SLF001
    assert trained_categories == {"profile-a", "profile-c"}

    service.import_classifier_model_artifact(
        filename="classifier.json",
        content=_classifier_artifact("profile-a", "profile-b"),
    )
    prediction = service.predict_classifier(ClassifierPredictRequest(question="任意の質問"))
    assert {candidate.category for candidate in prediction.candidates} == {"profile-a", "profile-b"}
    recommendation = service.recommend_profile(
        ProfileRecommendationRequest(question="任意の質問"),
        allowed_profile_ids={"profile-a"},
    )
    assert recommendation.recommendation_source == "classifier"
    assert set(recommendation.category_scores) == {"profile-a"}


def test_oracle_sync_job_routes_enforce_profile_access(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    service = _ProfileSyncService()
    monkeypatch.setattr(profile_sync_module, "profile_sync_service", service)
    request = _request({"profile-a"})

    with pytest.raises(HTTPException) as get_exc:
        nl2sql_router.get_profile_oracle_sync_job("sync-b", request)
    with pytest.raises(HTTPException) as retry_exc:
        nl2sql_router.retry_profile_oracle_sync_job("sync-b", request)

    assert get_exc.value.status_code == 403
    assert retry_exc.value.status_code == 403
    assert service.retried == []


def test_ontology_profile_recommendation_and_session_routes_enforce_profile_access(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    captured_allowed: set[str] | None = None

    def recommend_profiles(
        request: ontology_router.OntologyProfileRecommendationRequest,
        *,
        allowed_profile_ids: set[str] | None = None,
    ) -> ProfileRecommendation:
        del request
        nonlocal captured_allowed
        captured_allowed = allowed_profile_ids
        now = utc_now()
        return ProfileRecommendation(
            id="recommendation-1",
            question_hash="a" * 64,
            ontology_revision_id="revision-1",
            candidates=[
                ProfileRecommendationCandidateV2(
                    profile_id="profile-a",
                    profile_name="許可プロファイル",
                    ontology_revision_id="revision-1",
                    score=1.0,
                )
            ],
            created_at=now,
            expires_at=now,
        )

    monkeypatch.setattr(ontology_router.ontology_runtime, "recommend_profiles", recommend_profiles)
    request = _request({"profile-a"})

    ontology_router.recommend_ontology_profile(
        ontology_router.OntologyProfileRecommendationRequest(question="売上は?"),
        request,
    )
    assert captured_allowed == {"profile-a"}

    with pytest.raises(HTTPException) as confirm_exc:
        ontology_router.confirm_ontology_profile_recommendation(
            "recommendation-1",
            ontology_router.ProfileRecommendationConfirmationRequest(
                selected_profile_id="profile-b",
                selected_revision_id="revision-1",
            ),
            request,
        )
    with pytest.raises(HTTPException) as session_exc:
        ontology_router.create_query_session(
            ontology_router.QuerySessionApiCreate(question="売上は?", profile_id="profile-b"),
            request,
            idempotency_key="query-session-key",
        )

    assert confirm_exc.value.status_code == 403
    assert session_exc.value.status_code == 403
